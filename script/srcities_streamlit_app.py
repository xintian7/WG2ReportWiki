#!/usr/bin/env python3
"""Streamlit UI for browsing SRCities SPM Markdown content by major sections."""

from __future__ import annotations

import ast
import base64
import hmac
import html
import ipaddress
from io import BytesIO
import json
import logging
import os
from pathlib import Path
import re
from zipfile import BadZipFile, ZipFile
from collections.abc import Mapping

from cryptography.fernet import InvalidToken
from dotenv import load_dotenv
from encrypt_srsod import get_fernet
from markdown_it import MarkdownIt
from openpyxl import load_workbook
import plotly.graph_objects as go
import streamlit as st

LOGGER = logging.getLogger(__name__)

INLINE_TERM_TEXT = st.components.v2.component(
    "inline_glossary_terms",
    html="<div id='term-text'></div>",
    css="""
    #term-text {
        color: var(--st-text-color);
        font-family: var(--st-font);
        line-height: 1.55;
    }
    #term-text p { margin: 0 0 0.55rem; }
    #term-text a {
        color: #1a73e8;
        cursor: pointer;
        text-decoration: underline;
        text-decoration-thickness: 1px;
    }
    #term-text a:hover { color: #0b57d0; }
    #term-text .glossary-term-row { margin: 0 0 0.2rem; }
    #term-text .glossary-term-link { color: inherit; text-decoration: none; }
    #term-text .glossary-term-link:hover { color: inherit; text-decoration: underline; }
    #term-text .glossary-term-bullet { color: #000000; }
    #term-text .glossary-term-name { color: #000000; }
    #term-text .glossary-term-count { color: #7a4b2a; }
    #term-text .glossary-term-source { color: #0076a8; }
    #term-text .glossary-term-issue {
        color: #c62828;
        font-weight: 700;
        margin-left: 0.25rem;
    }
    #term-text .glossary-term-disabled { opacity: 0.45; }
    #term-text .glossary-term-disabled .glossary-term-bullet { color: #888888; }
    #term-text .glossary-term-disabled .glossary-term-name { color: #888888; }
    #term-text .glossary-term-disabled .glossary-term-count { color: #888888; }
    #term-text .glossary-term-disabled .glossary-term-source { color: #888888; }
    """,
    js="""
    export default function(component) {
        const { data, parentElement, setTriggerValue } = component;
        const root = parentElement.querySelector('#term-text');
        if (!root) return;

        if (root.innerHTML !== data.html) root.innerHTML = data.html;

        const onClick = (event) => {
            const termLink = event.target.closest('a[data-term]');
            const statementLink = event.target.closest('a[data-statement]');
            if (!termLink && !statementLink) return;
            event.preventDefault();
            event.stopPropagation();
            if (termLink) setTriggerValue('clicked', termLink.dataset.term);
            else setTriggerValue('clicked_statement', statementLink.dataset.statement);
        };

        root.addEventListener('click', onClick);
        return () => root.removeEventListener('click', onClick);
    }
    """,
)

GLOSSARY_NETWORK = st.components.v2.component(
    "glossary_network",
    html="<svg id='network' viewBox='0 0 1000 620' role='img'></svg>",
    css="""
    #network { width: 100%; height: 620px; background: transparent; border: none; }
    .edge { stroke: #90a0b8; stroke-opacity: .2; }
    .edge.connected { stroke: #5e42c6; stroke-opacity: .88; }
    .edge.dimmed { stroke-opacity: .06; }
    .node { fill: #5771ae; fill-opacity: .86; stroke: #ffffff; stroke-width: 1.5px; cursor: pointer; }
    .node.focused { fill: #1a73e8; stroke-width: 3px; }
    .node.dimmed { fill: #b9c0cd; opacity: .22; }
    .label { fill: var(--st-text-color); font: 12px var(--st-font); pointer-events: none; opacity: .62; }
    .label.dimmed { opacity: .08; }
    .label.focused { fill: #1a73e8; font-weight: 600; opacity: 1; }
    """,
    js="""
    export default function(component) {
        const { data, parentElement, setTriggerValue } = component;
        const svg = parentElement.querySelector('#network');
        if (!svg || !data) return;

        const width = 1000, height = 620;
        const centerX = width / 2, centerY = height / 2;
        const layoutRadius = Math.min(width, height) / 2 - 36;
        const nodeRadius = node => 4 + Math.min(8, Math.sqrt(node.count) * 2);
        const constrainToCircle = node => {
            const dx = node.x - centerX, dy = node.y - centerY;
            const distance = Math.hypot(dx, dy);
            const maxDistance = Math.max(0, layoutRadius - nodeRadius(node) - 10);
            if (distance > maxDistance) {
                node.x = centerX + dx / distance * maxDistance;
                node.y = centerY + dy / distance * maxDistance;
            }
        };
        const positions = parentElement._glossaryNetworkPositions ?? new Map();
        parentElement._glossaryNetworkPositions = positions;
        if (parentElement._glossaryNetworkLayoutVersion !== 'circular-v1') {
            positions.clear();
            parentElement._glossaryNetworkLayoutVersion = 'circular-v1';
        }
        const nodes = data.nodes.map((node, index) => ({
            ...node,
            x: positions.get(node.id)?.x ?? centerX + Math.cos(index * 2.399) * Math.sqrt((index + .5) / data.nodes.length) * (layoutRadius - 28),
            y: positions.get(node.id)?.y ?? centerY + Math.sin(index * 2.399) * Math.sqrt((index + .5) / data.nodes.length) * (layoutRadius - 28),
        }));
        const byId = new Map(nodes.map(node => [node.id, node]));
        const edges = data.edges.map(edge => ({ ...edge, source: byId.get(edge.source), target: byId.get(edge.target) }));
        const needsInitialLayout = positions.size !== nodes.length || nodes.some(node => !positions.has(node.id));

        if (needsInitialLayout) for (let iteration = 0; iteration < 180; iteration++) {
            for (let left = 0; left < nodes.length; left++) {
                for (let right = left + 1; right < nodes.length; right++) {
                    const first = nodes[left], second = nodes[right];
                    let dx = second.x - first.x, dy = second.y - first.y;
                    const distance = Math.max(12, Math.hypot(dx, dy));
                    const force = 3200 / (distance * distance);
                    dx /= distance; dy /= distance;
                    first.x -= dx * force; first.y -= dy * force;
                    second.x += dx * force; second.y += dy * force;
                }
            }
            edges.forEach(edge => {
                let dx = edge.target.x - edge.source.x, dy = edge.target.y - edge.source.y;
                const distance = Math.max(1, Math.hypot(dx, dy));
                const force = (distance - 260) * 0.009 * Math.min(edge.weight, 4);
                dx /= distance; dy /= distance;
                edge.source.x += dx * force; edge.source.y += dy * force;
                edge.target.x -= dx * force; edge.target.y -= dy * force;
            });
            nodes.forEach(node => {
                node.x += (centerX - node.x) * 0.003;
                node.y += (centerY - node.y) * 0.003;
                constrainToCircle(node);
            });
        }
        nodes.forEach(node => positions.set(node.id, { x: node.x, y: node.y }));

        let scale = 1, offsetX = 0, offsetY = 0, dragStart = null, nodeDrag = null;
        let hoveredNode = null, suppressClick = false;
        const svgPoint = event => {
            const rect = svg.getBoundingClientRect();
            return {
                x: (event.clientX - rect.left) * width / rect.width,
                y: (event.clientY - rect.top) * height / rect.height,
            };
        };
        const render = () => {
            const graph = document.createElementNS('http://www.w3.org/2000/svg', 'g');
            graph.setAttribute('transform', `translate(${offsetX} ${offsetY}) scale(${scale})`);
            edges.forEach(edge => {
                const line = document.createElementNS('http://www.w3.org/2000/svg', 'line');
                const isConnected = hoveredNode && (edge.source.id === hoveredNode || edge.target.id === hoveredNode);
                line.setAttribute('class', `edge${hoveredNode ? (isConnected ? ' connected' : ' dimmed') : ''}`);
                line.setAttribute('x1', edge.source.x); line.setAttribute('y1', edge.source.y);
                line.setAttribute('x2', edge.target.x); line.setAttribute('y2', edge.target.y);
                line.setAttribute('stroke-width', Math.min(3.5, 0.75 + edge.weight * 0.45));
                graph.appendChild(line);
            });
            nodes.forEach(node => {
                const group = document.createElementNS('http://www.w3.org/2000/svg', 'g');
                const circle = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
                const radius = 4 + Math.min(8, Math.sqrt(node.count) * 2);
                const isNeighbor = hoveredNode && edges.some(edge =>
                    (edge.source.id === hoveredNode && edge.target.id === node.id) ||
                    (edge.target.id === hoveredNode && edge.source.id === node.id)
                );
                const nodeClass = hoveredNode
                    ? `node${node.id === hoveredNode ? ' focused' : isNeighbor ? ' focused' : ' dimmed'}`
                    : 'node';
                circle.setAttribute('class', nodeClass); circle.setAttribute('cx', node.x); circle.setAttribute('cy', node.y); circle.setAttribute('r', radius);
                const title = document.createElementNS('http://www.w3.org/2000/svg', 'title'); title.textContent = `${node.label} [${node.count}]`;
                circle.appendChild(title);
                circle.addEventListener('mouseenter', () => { hoveredNode = node.id; render(); });
                circle.addEventListener('mouseleave', () => { hoveredNode = null; render(); });
                circle.addEventListener('pointerdown', event => {
                    event.stopPropagation();
                    const point = svgPoint(event);
                    nodeDrag = { node, x: point.x, y: point.y, nodeX: node.x, nodeY: node.y, moved: false };
                    svg.setPointerCapture(event.pointerId);
                });
                circle.addEventListener('click', event => {
                    event.stopPropagation();
                    if (!suppressClick) setTriggerValue('clicked', node.label);
                    suppressClick = false;
                });
                const label = document.createElementNS('http://www.w3.org/2000/svg', 'text');
                const labelClass = hoveredNode
                    ? `label${node.id === hoveredNode || isNeighbor ? ' focused' : ' dimmed'}`
                    : 'label';
                label.setAttribute('class', labelClass);
                label.setAttribute('x', node.x + radius + 4); label.setAttribute('y', node.y + 4); label.textContent = node.label;
                group.appendChild(circle); group.appendChild(label); graph.appendChild(group);
            });
            svg.replaceChildren(graph);
        };
        const clearHighlight = () => {
            hoveredNode = null;
            render();
        };
        svg.onwheel = event => { event.preventDefault(); const factor = event.deltaY < 0 ? 1.12 : 0.89; scale = Math.max(.45, Math.min(3, scale * factor)); render(); };
        svg.onpointerdown = event => { dragStart = { x: event.clientX, y: event.clientY, offsetX, offsetY, moved: false }; svg.setPointerCapture(event.pointerId); };
        svg.onpointermove = event => {
            if (nodeDrag) {
                const point = svgPoint(event);
                const dx = (point.x - nodeDrag.x) / scale, dy = (point.y - nodeDrag.y) / scale;
                nodeDrag.node.x = nodeDrag.nodeX + dx;
                nodeDrag.node.y = nodeDrag.nodeY + dy;
                positions.set(nodeDrag.node.id, { x: nodeDrag.node.x, y: nodeDrag.node.y });
                nodeDrag.moved = nodeDrag.moved || Math.abs(dx) > 2 || Math.abs(dy) > 2;
                render();
                return;
            }
            if (!dragStart) return;
            offsetX = dragStart.offsetX + (event.clientX - dragStart.x); offsetY = dragStart.offsetY + (event.clientY - dragStart.y); render();
            dragStart.moved = dragStart.moved || Math.abs(event.clientX - dragStart.x) > 2 || Math.abs(event.clientY - dragStart.y) > 2;
        };
        svg.onpointerup = () => {
            if (nodeDrag) suppressClick = nodeDrag.moved;
            else if (dragStart && !dragStart.moved) clearHighlight();
            nodeDrag = null;
            dragStart = null;
        };
        render();
    }
    """,
)

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
USER_PASSWORDS_ENV_NAME = "user_pwd"
load_dotenv(REPO_ROOT / ".env", override=False)
ENCRYPTED_REPORT_PATH = REPO_ROOT / "data" / "encrypted" / "SRSOD.enc"
USER_GUIDE_PATH = REPO_ROOT / "data" / "user_guide.html"
USER_GUIDE_SETTINGS_IMAGE_PATH = REPO_ROOT / "data" / "iamges4UG" / "image1.png"
USER_GUIDE_GLOSSARY_IMAGE_PATH = REPO_ROOT / "data" / "iamges4UG" / "image2.png"
USER_GUIDE_LLM_SUMMARY_IMAGE_PATH = REPO_ROOT / "data" / "iamges4UG" / "image3.png"
REPORT_ARCHIVE_NAME = "data/report/SRCities_FOD_SPM_Final.md"
TERM_USAGE_SUMMARY_PROMPT_ARCHIVE_NAME = "data/prompt/llm_term_usage_summary_prompt.md"
TERM_USAGE_SUMMARIES_ARCHIVE_NAME = "data/analysis/term_usage_summaries.json"
GLOSSARY_ARCHIVE_MEMBER_NAMES = (
    "data/Glossary/AR6FGD_Glossary.xlsx",
    "data/Glossary/AR7SOD_Glossary.xlsx",
)
GLOSSARY_NETWORK_ARCHIVE_MEMBER_NAME = "data/network/SRCities_glossary_network.cypher"
EXECUTIVE_SUMMARY_ARCHIVE_MEMBER_NAMES = {
    f"Chapter {chapter}": f"data/ES/SRCities_SOD_Ch{chapter:02d}_Final_executive_summary.md"
    for chapter in range(1, 6)
}
GLOSSARY_OVERVIEW_LABEL = "Glossary Overview"
GLOSSARY_NETWORK_LABEL = "Glossary Network (under development)"
SANKEY_DIAGRAM_LABEL = "Sankey diagram (under development)"
SETTINGS_ACCESS_LABEL = "Access"
USER_GUIDE_LABEL = "User Guide"
LLM_SUMMARY_EXCLUDED_TERM_KEYS = frozenset({"confidence"})
AR6_SOURCE_LABEL = "AR6"
LEGACY_AR6_SOURCE_LABEL = "AR6-FGD"

SECTION_LABELS = [
    "Author Team",
    "Introduction",
    "Section A",
    "Section B",
    "Section C",
    "Section D",
    GLOSSARY_OVERVIEW_LABEL,
    GLOSSARY_NETWORK_LABEL,
    SANKEY_DIAGRAM_LABEL,
]
SPM_SECTION_LABELS = SECTION_LABELS[:6]
GLOSSARY_SECTION_LABELS = SECTION_LABELS[6:]
NAVIGATION_GROUP_KEYS = {
    "user_guide": "user_guide_section",
    "executive_summaries": "executive_summary_section",
    "spm": "spm_nav_section",
    "glossary": "glossary_nav_section",
}

STATEMENT_NUMBER_RE = re.compile(r"^([A-D]\d+(?:\.\d+)+)\s+(.*)$")
FIRST_LEVEL_STATEMENT_RE = re.compile(r"^([A-D]\d)\s+(.*)$")
FIGURE_CAPTION_RE = re.compile(r"^Figure\s+(SPM\.\d+):\s+(.*)$")
EXECUTIVE_SUMMARY_ITEM_RE = re.compile(r"^(?:#{1,6}\s+)?\[(ES\d+\.\d+(?:\.\d+)?)\]\s+(.*)$")
NUMERIC_REFERENCE_RE = re.compile(r"^\d+(?:\.\d+)+$")
REFERENCE_BLOCK_RE = re.compile(r"\{([^}]*)\}")
POTENTIAL_ISSUES_SECTION_RE = re.compile(
    r"^###\s+Potential issues(?:\s+(?:needing|for))?\s+substantive review[^\n]*\n"
    r"(?P<content>.*?)(?=\n###\s|\Z)",
    re.IGNORECASE | re.MULTILINE | re.DOTALL,
)
NO_POTENTIAL_ISSUE_RE = re.compile(r"^-?\s*none identified\b", re.IGNORECASE)
GlossaryDefinition = tuple[str, str, str, str]
Glossary = dict[str, list[GlossaryDefinition]]
ExecutiveSummaries = dict[str, str]
StatementOccurrence = tuple[str, str, str]
TermUsageSummaries = dict[str, str]
MARKDOWN_RENDERER = MarkdownIt("commonmark", {"html": False})
SUMMARY_MARKDOWN_RENDERER = MarkdownIt("commonmark", {"html": False}).enable("table")


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def normalize_source_label(source: str) -> str:
    """Map legacy glossary source labels to their current display names."""
    return AR6_SOURCE_LABEL if source == LEGACY_AR6_SOURCE_LABEL else source


def llm_summary_not_applied_message(term: str) -> str:
    """Return the notice shown for terms excluded from LLM analysis."""
    return f"LLM is not applied to the term {term}."


def summary_has_potential_issue(summary: str) -> bool:
    """Return whether a bundled LLM summary reports a substantive-review issue."""
    match = POTENTIAL_ISSUES_SECTION_RE.search(summary)
    if not match:
        return False
    issue_text = re.sub(r"[*_`]", "", match.group("content")).strip()
    return NO_POTENTIAL_ISSUE_RE.match(issue_text) is None


def build_glossary_pattern(glossary: Glossary) -> re.Pattern[str] | None:
    """Build one regex pattern to find all glossary terms (longest first)."""
    if not glossary:
        return None

    terms = sorted(glossary.keys(), key=len, reverse=True)
    if not terms:
        return None

    pattern = "|".join(re.escape(term) for term in terms)
    if not pattern:
        return None

    return re.compile(pattern, re.IGNORECASE)


def term_anchor_id(term_key: str) -> str:
    """Build a stable in-page anchor id for a glossary term key."""
    slug = re.sub(r"[^a-z0-9]+", "-", term_key.lower()).strip("-")
    return f"term-{slug or 'item'}"


def linkify_glossary_terms(
    text: str,
    glossary: Glossary,
    term_pattern: re.Pattern[str] | None,
) -> str:
    """Render text as safe HTML with inline glossary-term links."""
    if not glossary or term_pattern is None:
        return html.escape(text)

    output_parts: list[str] = []
    last_end = 0

    def is_word_char(ch: str) -> bool:
        return ch.isalnum() or ch == "_"

    for match in term_pattern.finditer(text):
        start, end = match.span()
        matched_text = match.group(0)
        key = normalize_text(matched_text).casefold()
        glossary_entry = glossary.get(key)

        if not glossary_entry:
            continue

        # Avoid linking fragments inside larger words, e.g., "risk" in "risks".
        if matched_text and matched_text[0].isalnum() and start > 0 and is_word_char(text[start - 1]):
            continue
        if matched_text and matched_text[-1].isalnum() and end < len(text) and is_word_char(text[end]):
            continue

        output_parts.append(html.escape(text[last_end:start]))

        safe_label = html.escape(matched_text)
        safe_term = html.escape(glossary_entry[0][0], quote=True)
        output_parts.append(f'<a href="#" data-term="{safe_term}">{safe_label}</a>')

        last_end = end

    output_parts.append(html.escape(text[last_end:]))
    return "".join(output_parts)


def linkify_glossary_html(
    rendered_html: str,
    glossary: Glossary,
    term_pattern: re.Pattern[str] | None,
) -> str:
    """Add glossary links to Markdown-rendered text without changing its HTML markup."""
    output_parts: list[str] = []
    anchor_depth = 0

    for fragment in re.split(r"(<[^>]+>)", rendered_html):
        tag_match = re.match(r"</?([a-zA-Z][a-zA-Z0-9]*)", fragment)
        if not tag_match:
            output_parts.append(
                fragment if anchor_depth else linkify_glossary_terms(html.unescape(fragment), glossary, term_pattern)
            )
            continue

        tag_name = tag_match.group(1).lower()
        if tag_name == "a" and fragment.startswith("</"):
            anchor_depth -= 1
        output_parts.append(fragment)
        if tag_name == "a" and not fragment.startswith("</") and not fragment.endswith("/>"):
            anchor_depth += 1

    return "".join(output_parts)


def linkify_summary_statement_references(
    rendered_html: str,
    matches: list[StatementOccurrence],
) -> str:
    """Link known statement identifiers in an LLM summary to their source text."""
    statement_indices = {
        normalize_text(statement_number).casefold(): index
        for index, (statement_number, _, _) in enumerate(matches)
    }
    if not statement_indices:
        return rendered_html

    identifiers = sorted(statement_indices, key=len, reverse=True)
    identifier_pattern = re.compile(
        rf"(?<![A-Za-z0-9.-])(?P<identifier>{'|'.join(re.escape(identifier) for identifier in identifiers)})"
        r"(?![A-Za-z0-9.-])",
        re.IGNORECASE,
    )

    def linkify_text(text: str) -> str:
        parts: list[str] = []
        last_end = 0
        for match in identifier_pattern.finditer(text):
            identifier = match.group("identifier")
            index = statement_indices[identifier.casefold()]
            parts.append(html.escape(text[last_end:match.start()]))
            parts.append(
                f"<a href='#' data-statement='{index}' title='Open full statement'>{html.escape(identifier)}</a>"
            )
            last_end = match.end()
        parts.append(html.escape(text[last_end:]))
        return "".join(parts)

    output_parts: list[str] = []
    protected_depth = 0
    protected_tags = {"a", "code", "pre"}
    for fragment in re.split(r"(<[^>]+>)", rendered_html):
        tag_match = re.match(r"</?([a-zA-Z][a-zA-Z0-9]*)", fragment)
        if not tag_match:
            output_parts.append(fragment if protected_depth else linkify_text(html.unescape(fragment)))
            continue

        tag_name = tag_match.group(1).lower()
        if tag_name in protected_tags and fragment.startswith("</"):
            protected_depth -= 1
        output_parts.append(fragment)
        if tag_name in protected_tags and not fragment.startswith("</") and not fragment.endswith("/>"):
            protected_depth += 1

    return "".join(output_parts)


def markdown_to_plain_text(markdown_text: str) -> str:
    """Convert a Markdown fragment to display text for a Terms-table row."""
    rendered_html = MARKDOWN_RENDERER.render(markdown_text)
    return normalize_text(html.unescape(re.sub(r"<[^>]+>", "", rendered_html)))


def markdown_table_cell(text: str) -> str:
    """Escape a value for the Markdown table supplied to the LLM prompt."""
    return normalize_text(text).replace("|", r"\|")


def definition_for_source(definitions: list[GlossaryDefinition], source: str) -> str:
    """Return all available definitions from one glossary source."""
    source_definitions = [
        definition for _, definition, _, definition_source in definitions if definition_source == source
    ]
    return "\n\n".join(source_definitions) or "Definition not available"


def glossary_source_label(definitions: list[GlossaryDefinition]) -> str:
    """Format glossary definition sources in their established display order."""
    source_order = ("SRCities-SOD", AR6_SOURCE_LABEL)
    available_sources = {source for _, _, _, source in definitions}
    return " + ".join(source for source in source_order if source in available_sources)


def glossary_parent_label(definitions: list[GlossaryDefinition]) -> str:
    """Format the unique parent term values available for a glossary entry."""
    parents = list(dict.fromkeys(parent for _, _, parent, _ in definitions if parent))
    return "; ".join(parents)


def build_term_usage_summary_prompt(
    prompt_template: str,
    term: str,
    definitions: list[GlossaryDefinition],
    matches: list[StatementOccurrence],
) -> str:
    """Fill the approved term-usage prompt with glossary definitions and occurrence rows."""
    if not prompt_template.strip():
        raise ValueError("LLM summary prompt is empty.")

    occurrence_rows = [
        f"| {markdown_table_cell(statement_number)} | {markdown_table_cell(sentence)} |"
        for statement_number, sentence, _ in matches
    ]
    statement_sentence_table = "\n".join(
        [
            f"Occurrence count: {len(matches)}",
            "",
            "| Statement | Sentence |",
            "| --- | --- |",
            *occurrence_rows,
        ]
    )
    replacements = {
        "{term}": term,
        "{srcities_sod_definition_or_not_available}": definition_for_source(definitions, "SRCities-SOD"),
        "{ar6_fgd_definition_or_not_available}": definition_for_source(definitions, AR6_SOURCE_LABEL),
        "{statement_sentence_table}": statement_sentence_table,
    }
    prompt = prompt_template
    for placeholder, value in replacements.items():
        prompt = prompt.replace(placeholder, value)
    return prompt


def term_occurrences(
    term: str,
    current_section: str,
    sections: dict[str, list[str]],
    executive_summaries: ExecutiveSummaries | None = None,
) -> list[StatementOccurrence]:
    """Find matching SPM sentences, then numbered Executive Summary sentences."""
    if not term:
        return []

    pattern = re.compile(rf"(?<!\w){re.escape(term)}(?!\w)", re.IGNORECASE)
    matches: list[StatementOccurrence] = []

    for section_name, lines in sections.items():
        if section_name == current_section:
            continue
        for line in lines:
            number_match = STATEMENT_NUMBER_RE.match(line)
            first_level_match = FIRST_LEVEL_STATEMENT_RE.match(line)
            figure_match = FIGURE_CAPTION_RE.match(line)
            if figure_match and section_name.startswith("Section "):
                section_identifier = section_name.removeprefix("Section ")
                statement_number = f"{section_identifier}-Figure {figure_match.group(1)}"
                statement_text = figure_match.group(2)
                excerpts = [statement_text]
            elif number_match:
                statement_number, statement_text = number_match.groups()
                excerpts = re.split(r"(?<=[.!?])\s+", statement_text)
            elif first_level_match:
                statement_number, statement_text = first_level_match.groups()
                excerpts = re.split(r"(?<=[.!?])\s+", statement_text)
            else:
                statement_number, statement_text = section_name, line
                excerpts = re.split(r"(?<=[.!?])\s+", statement_text)

            for sentence in excerpts:
                if pattern.search(sentence):
                    matches.append((statement_number, sentence.strip(), statement_text))

    if executive_summaries:
        for summary_text in executive_summaries.values():
            for raw_line in summary_text.splitlines():
                item_match = EXECUTIVE_SUMMARY_ITEM_RE.match(raw_line.strip())
                if not item_match:
                    continue

                statement_number, markdown_text = item_match.groups()
                statement_text = markdown_to_plain_text(markdown_text)
                for sentence in re.split(r"(?<=[.!?])\s+", statement_text):
                    if pattern.search(sentence):
                        matches.append((statement_number, sentence.strip(), markdown_text))

    return matches


def glossary_usage_sentences(
    sections: dict[str, list[str]],
    executive_summaries: ExecutiveSummaries,
) -> list[str]:
    """Extract the same SPM and ES sentence units used by term occurrences."""
    sentences: list[str] = []

    for section_name, lines in sections.items():
        for line in lines:
            number_match = STATEMENT_NUMBER_RE.match(line)
            first_level_match = FIRST_LEVEL_STATEMENT_RE.match(line)
            figure_match = FIGURE_CAPTION_RE.match(line)
            if figure_match and section_name.startswith("Section "):
                statement_text = figure_match.group(2)
                excerpts = [statement_text]
            elif number_match:
                _, statement_text = number_match.groups()
                excerpts = re.split(r"(?<=[.!?])\s+", statement_text)
            elif first_level_match:
                _, statement_text = first_level_match.groups()
                excerpts = re.split(r"(?<=[.!?])\s+", statement_text)
            else:
                excerpts = re.split(r"(?<=[.!?])\s+", line)
            sentences.extend(excerpts)

    for summary_text in executive_summaries.values():
        for raw_line in summary_text.splitlines():
            item_match = EXECUTIVE_SUMMARY_ITEM_RE.match(raw_line.strip())
            if not item_match:
                continue
            _, markdown_text = item_match.groups()
            sentences.extend(re.split(r"(?<=[.!?])\s+", markdown_to_plain_text(markdown_text)))

    return sentences


def glossary_usage_counts(
    glossary: Glossary,
    sections: dict[str, list[str]],
    executive_summaries: ExecutiveSummaries,
) -> dict[str, int]:
    """Count each glossary term once per matching SPM or ES sentence."""
    terms_by_initial: dict[str, list[str]] = {}
    for term_key in glossary:
        terms_by_initial.setdefault(term_key[0], []).append(term_key)

    counts = {term_key: 0 for term_key in glossary}

    def is_word_character(character: str) -> bool:
        return character.isalnum() or character == "_"

    for sentence in glossary_usage_sentences(sections, executive_summaries):
        normalized_sentence = sentence.casefold()
        matched_terms: set[str] = set()
        for start, character in enumerate(normalized_sentence):
            if start and is_word_character(normalized_sentence[start - 1]):
                continue
            for term_key in terms_by_initial.get(character, []):
                end = start + len(term_key)
                if normalized_sentence.startswith(term_key, start) and (
                    end == len(normalized_sentence) or not is_word_character(normalized_sentence[end])
                ):
                    matched_terms.add(term_key)
        for term_key in matched_terms:
            counts[term_key] += 1

    return counts


def statement_occurrence_from_trigger(
    clicked_index: str | int,
    matches: list[StatementOccurrence],
) -> StatementOccurrence | None:
    """Resolve an inline-component statement click to its full source occurrence."""
    try:
        index = int(clicked_index)
    except (TypeError, ValueError):
        return None

    return matches[index] if 0 <= index < len(matches) else None


def highlight_term(text: str, term: str) -> str:
    """Highlight matched term in a text snippet."""
    pattern = re.compile(rf"(?<!\w){re.escape(term)}(?!\w)", re.IGNORECASE)
    escaped = html.escape(text)
    return pattern.sub(lambda m: f"<mark>{html.escape(m.group(0))}</mark>", escaped)


def terms_in_lines(lines: list[str], glossary: Glossary, term_pattern: re.Pattern[str] | None) -> list[str]:
    """Collect glossary term keys present in given lines."""
    if term_pattern is None:
        return []

    keys: set[str] = set()

    def is_word_char(ch: str) -> bool:
        return ch.isalnum() or ch == "_"

    for line in lines:
        for match in term_pattern.finditer(line):
            start, end = match.span()
            matched_text = match.group(0)
            key = normalize_text(matched_text).casefold()
            if key not in glossary:
                continue
            if matched_text and matched_text[0].isalnum() and start > 0 and is_word_char(line[start - 1]):
                continue
            if matched_text and matched_text[-1].isalnum() and end < len(line) and is_word_char(line[end]):
                continue
            keys.add(key)

    return sorted(keys)


def terms_in_line(line: str, glossary: Glossary, term_pattern: re.Pattern[str] | None) -> list[str]:
    """Collect glossary term keys found in one line, preserving first-seen order."""
    if term_pattern is None:
        return []

    keys: list[str] = []
    seen: set[str] = set()

    def is_word_char(ch: str) -> bool:
        return ch.isalnum() or ch == "_"

    for match in term_pattern.finditer(line):
        start, end = match.span()
        matched_text = match.group(0)
        key = normalize_text(matched_text).casefold()
        if key not in glossary or key in seen:
            continue
        if matched_text and matched_text[0].isalnum() and start > 0 and is_word_char(line[start - 1]):
            continue
        if matched_text and matched_text[-1].isalnum() and end < len(line) and is_word_char(line[end]):
            continue
        seen.add(key)
        keys.append(key)

    return keys


def statement_references(sections: dict[str, list[str]]) -> list[tuple[str, str]]:
    """Map each numbered SPM finding to the numeric source references it cites."""
    links: list[tuple[str, str]] = []

    for lines in sections.values():
        for line in lines:
            statement_match = STATEMENT_NUMBER_RE.match(line)
            if not statement_match:
                continue

            statement_number, statement_text = statement_match.groups()
            references: list[str] = []
            for reference_block in REFERENCE_BLOCK_RE.findall(statement_text):
                for reference in re.split(r"[;,]", reference_block):
                    reference = normalize_text(reference)
                    if NUMERIC_REFERENCE_RE.fullmatch(reference) and reference not in references:
                        references.append(reference)

            links.extend((statement_number, reference) for reference in references)

    return links


def render_sankey_diagram(sections: dict[str, list[str]]) -> None:
    """Render SPM findings and their cited report sections as a Sankey diagram."""
    st.subheader(SANKEY_DIAGRAM_LABEL)
    links = statement_references(sections)
    if not links:
        st.info("No numbered finding-to-reference links were found.")
        return

    section_options = ["All sections", "Section A", "Section B", "Section C", "Section D"]
    selected_section = st.selectbox("SPM findings", section_options, key="sankey_section")
    if selected_section != "All sections":
        section_prefix = selected_section[-1]
        links = [(statement, reference) for statement, reference in links if statement.startswith(section_prefix)]

    statement_nodes = sorted(
        {statement for statement, _ in links},
        key=lambda statement: (statement[0], tuple(int(part) for part in statement[1:].split("."))),
    )
    reference_nodes = sorted(
        {reference for _, reference in links},
        key=lambda reference: tuple(int(part) for part in reference.split(".")),
    )
    labels = statement_nodes + reference_nodes
    node_index = {label: index for index, label in enumerate(labels)}

    figure = go.Figure(
        go.Sankey(
            arrangement="snap",
            node={
                "label": labels,
                "x": [0.02] * len(statement_nodes) + [0.98] * len(reference_nodes),
                "pad": 18,
                "thickness": 20,
                "color": ["#176B87"] * len(statement_nodes) + ["#C9672A"] * len(reference_nodes),
                "line": {"color": "#ffffff", "width": 1.5},
            },
            link={
                "source": [node_index[statement] for statement, _ in links],
                "target": [node_index[reference] for _, reference in links],
                "value": [1] * len(links),
                "color": "rgba(49, 106, 137, 0.32)",
                "hovertemplate": "%{source.label} cites %{target.label}<extra></extra>",
            },
        )
    )
    figure.update_layout(
        height=max(760, min(2400, 24 * max(len(statement_nodes), len(reference_nodes)))),
        margin={"l": 20, "r": 20, "t": 70, "b": 20},
        font={"size": 12},
        annotations=[
            {
                "x": 0.02,
                "y": 1.08,
                "xref": "paper",
                "yref": "paper",
                "text": "<b>SPM findings</b><br>Numbered statements",
                "showarrow": False,
                "xanchor": "left",
                "align": "left",
            },
            {
                "x": 0.98,
                "y": 1.08,
                "xref": "paper",
                "yref": "paper",
                "text": "<b>Cited report sections</b><br>Numeric evidence references",
                "showarrow": False,
                "xanchor": "right",
                "align": "right",
            },
        ],
    )
    st.plotly_chart(figure, width="stretch", key="sankey_diagram")


def parse_glossary_network(cypher_text: str) -> dict[str, list[dict[str, str | int]]]:
    """Parse node and edge metadata embedded in the generated Cypher dataset."""
    nodes: list[dict[str, str | int]] = []
    edges: list[dict[str, str | int]] = []
    for line in cypher_text.splitlines():
        if line.startswith("// NODE "):
            nodes.append(json.loads(line.removeprefix("// NODE ")))
        elif line.startswith("// EDGE "):
            edges.append(json.loads(line.removeprefix("// EDGE ")))
    return {"nodes": nodes, "edges": edges}


def get_selected_term() -> str:
    """Get selected term from session state (initialized from URL when available)."""
    value = st.session_state.get("selected_term", "")
    return normalize_text(str(value))


def select_navigation_group(group: str) -> None:
    """Activate one navigation group and clear all other group selections."""
    st.session_state.active_navigation_group = group
    for navigation_group, key in NAVIGATION_GROUP_KEYS.items():
        if navigation_group != group:
            st.session_state[key] = None


@st.dialog("Statement text", width="large")
def show_statement_dialog(statement_number: str, statement_text: str) -> None:
    """Display the full SPM or Executive Summary text for a selected statement."""
    st.markdown(f"**{html.escape(statement_number)}**")
    st.markdown(statement_text)


def render_term_details(
    selected_term: str,
    current_section: str,
    sections: dict[str, list[str]],
    glossary: Glossary,
    executive_summaries: ExecutiveSummaries,
    term_usage_summaries: TermUsageSummaries,
    include_current_section: bool = False,
) -> None:
    """Render term details and matching sentence table."""
    st.subheader("Terms")
    if not selected_term:
        st.caption("Click a term in the left panel to show its definition and related texts.")
        return

    key = selected_term.casefold()
    if key not in glossary:
        st.warning(f"No glossary definition found for '{selected_term}'.")
        return

    definitions = glossary[key]
    term = definitions[0][0]
    excluded_section = "" if include_current_section else current_section
    matches = term_occurrences(term, excluded_section, sections, executive_summaries)
    parent = glossary_parent_label(definitions)
    sources = glossary_source_label(definitions)

    st.markdown(
        f"<span style='color:#000000;'><strong>{html.escape(term)}</strong></span> "
        f"<span style='color:#7a4b2a;'><strong>[{len(matches)}]</strong></span> "
        f"<span style='color:#0076a8;'><strong>[{html.escape(sources)}]</strong></span>",
        unsafe_allow_html=True,
    )
    if parent:
        st.markdown(f"Parent: {html.escape(parent)}")
    for _, definition, _, source in definitions:
        st.markdown(f"**Explanation in {html.escape(source)}**")
        st.write(definition)

    summary_clicked = st.button("LLM Summary", key=f"llm_summary_{key}")

    summary_state_key = f"llm_summary_result_{key}"
    if summary_clicked:
        summary = (
            llm_summary_not_applied_message(term)
            if key in LLM_SUMMARY_EXCLUDED_TERM_KEYS
            else term_usage_summaries.get(key, "")
        )
        if summary:
            st.session_state[summary_state_key] = summary
        else:
            st.session_state.pop(summary_state_key, None)
            st.info(
                "No precomputed LLM summary is available for this term. "
                "Run script/generate_term_usage_summaries.py and rebuild SRSOD.enc."
            )

    summary = st.session_state.get(summary_state_key, "")
    if summary:
        with st.expander("LLM summary", expanded=True):
            if key in LLM_SUMMARY_EXCLUDED_TERM_KEYS:
                st.info(summary)
            else:
                st.info(
                    "⚠️ This LLM-assisted summary, including the potential inconsistency flags, is provided only as "
                    "a reference. It may contain errors "
                    "and it is not intended to replace the human judgment for consistency checks."
                )
                summary_html = linkify_summary_statement_references(
                    SUMMARY_MARKDOWN_RENDERER.render(summary),
                    matches,
                )
                summary_result = INLINE_TERM_TEXT(
                    data={"html": summary_html},
                    key=f"llm_summary_content_{key}",
                    on_clicked_statement_change=lambda: None,
                )
                selected_summary_occurrence = statement_occurrence_from_trigger(
                    getattr(summary_result, "clicked_statement", ""),
                    matches,
                )
                if selected_summary_occurrence:
                    statement_number, _, statement_text = selected_summary_occurrence
                    show_statement_dialog(statement_number, statement_text)

    if not matches:
        st.info("No usage found in other sections.")
        return

    rows = [
        "<table style='width:100%; border-collapse:collapse;'>"
        "<thead><tr><th style='text-align:left; border-bottom:1px solid #ddd; padding:6px;'>Statement</th>"
        "<th style='text-align:left; border-bottom:1px solid #ddd; padding:6px;'>Sentence</th></tr></thead><tbody>"
    ]
    for index, (statement_number, sentence, _) in enumerate(matches):
        highlighted = highlight_term(sentence, term)
        rows.append(
            "<tr>"
            "<td style='vertical-align:top; border-bottom:1px solid #eee; padding:6px; white-space:nowrap;'>"
            f"<a href='#' data-statement='{index}' title='Open full statement'>{html.escape(statement_number)}</a>"
            "</td>"
            f"<td style='border-bottom:1px solid #eee; padding:6px;'>{highlighted}</td>"
            "</tr>"
        )
    rows.append("</tbody></table>")
    result = INLINE_TERM_TEXT(
        data={"html": "".join(rows)},
        key=f"term_occurrence_table_{key}",
        on_clicked_statement_change=lambda: None,
    )
    selected_occurrence = statement_occurrence_from_trigger(
        getattr(result, "clicked_statement", ""),
        matches,
    )
    if selected_occurrence:
        statement_number, _, statement_text = selected_occurrence
        show_statement_dialog(statement_number, statement_text)


def render_glossary(
    glossary: Glossary,
    sections: dict[str, list[str]],
    executive_summaries: ExecutiveSummaries,
    term_usage_summaries: TermUsageSummaries,
) -> None:
    """Render merged glossary terms with source labels and SPM/ES occurrence counts."""
    if not glossary:
        st.subheader(f"{GLOSSARY_OVERVIEW_LABEL} (0/0)")
        st.info("No glossary entries are available.")
        return

    usage_counts = glossary_usage_counts(glossary, sections, executive_summaries)
    used_term_count = sum(frequency > 0 for frequency in usage_counts.values())
    st.subheader(f"{GLOSSARY_OVERVIEW_LABEL} ({used_term_count}/{len(glossary)})")
    
    # Add search box
    search_term = st.text_input("Search for a term", placeholder="Type a term name...")
    search_query = search_term.replace("!", "").strip().casefold()
    search_potential_issues = "!" in search_term
    
    rows: list[str] = []
    for definitions in sorted(glossary.values(), key=lambda entries: entries[0][0].casefold()):
        term = definitions[0][0]
        term_key = term.casefold()
        has_potential_issue = (
            term_key not in LLM_SUMMARY_EXCLUDED_TERM_KEYS
            and summary_has_potential_issue(term_usage_summaries.get(term_key, ""))
        )
        if search_potential_issues and not has_potential_issue:
            continue
        if search_query and search_query not in term_key:
            continue
        frequency = usage_counts[term_key]
        sources = glossary_source_label(definitions)
        issue_indicator = (
            "<span class='glossary-term-issue' role='img' "
            "aria-label='Potential issue needing substantive review' "
            "title='Potential issue needing substantive review'>!</span>"
            if has_potential_issue
            else ""
        )
        label = (
            "<span class='glossary-term-bullet'>• </span>"
            f"<span class='glossary-term-name'>{html.escape(term)}</span> "
            f"<span class='glossary-term-count'>[{frequency}]</span> "
            f"<span class='glossary-term-source'>[{html.escape(sources)}]</span>{issue_indicator}"
        )
        if frequency:
            rows.append(
                "<p class='glossary-term-row'>"
                f"<a class='glossary-term-link' href='#' data-term='{html.escape(term, quote=True)}'>{label}</a>"
                "</p>"
            )
        else:
            rows.append(f"<p class='glossary-term-row glossary-term-disabled'>{label}</p>")

    result = INLINE_TERM_TEXT(
        data={"html": "".join(rows)},
        key="glossary_term_list",
        on_clicked_change=lambda: None,
    )
    clicked_term = getattr(result, "clicked", "")
    if clicked_term:
        st.session_state.selected_term = normalize_text(str(clicked_term))


def open_network_term_in_glossary() -> None:
    """Open the clicked network term in the Glossary view."""
    component_state = st.session_state.get("glossary_network_graph", {})
    clicked_term = component_state.get("clicked", "")
    if clicked_term:
        st.session_state.selected_term = normalize_text(str(clicked_term))
        st.session_state.active_navigation_group = "glossary"
        st.session_state.glossary_nav_section = GLOSSARY_OVERVIEW_LABEL


def render_glossary_network(graph: dict[str, list[dict[str, str | int]]]) -> None:
    """Render the interactive glossary co-occurrence network."""
    st.subheader(GLOSSARY_NETWORK_LABEL)
    if not graph["nodes"]:
        st.info("No glossary network data is available in the encrypted archive.")
        return

    counts = [int(node["count"]) for node in graph["nodes"]]
    lowest_count = min(counts)
    highest_count = max(counts)
    default_lowest_count = max(lowest_count, min(5, highest_count))
    term_count_column, connection_count_column = st.columns(2)
    with term_count_column:
        minimum_count, maximum_count = st.slider(
            "Term usage count",
            min_value=lowest_count,
            max_value=highest_count,
            value=(default_lowest_count, highest_count),
            step=1,
            key="glossary_network_count_range",
        )
    filtered_nodes = [
        node
        for node in graph["nodes"]
        if minimum_count <= int(node["count"]) <= maximum_count
    ]
    filtered_node_ids = {str(node["id"]) for node in filtered_nodes}
    connection_counts = [int(edge["weight"]) for edge in graph["edges"]]
    if connection_counts:
        lowest_connection_count = min(connection_counts)
        highest_connection_count = max(connection_counts)
        default_lowest_connection_count = max(
            lowest_connection_count,
            min(5, highest_connection_count),
        )
        with connection_count_column:
            minimum_connection_count, maximum_connection_count = st.slider(
                "Connection count",
                min_value=lowest_connection_count,
                max_value=highest_connection_count,
                value=(default_lowest_connection_count, highest_connection_count),
                step=1,
                key="glossary_network_connection_count_range",
            )
    else:
        minimum_connection_count = None
        maximum_connection_count = None
    filtered_edges = [
        edge
        for edge in graph["edges"]
        if (
            str(edge["source"]) in filtered_node_ids
            and str(edge["target"]) in filtered_node_ids
            and (
                minimum_connection_count is None
                or minimum_connection_count <= int(edge["weight"]) <= maximum_connection_count
            )
        )
    ]
    connected_node_ids = {
        str(edge[node_id])
        for edge in filtered_edges
        for node_id in ("source", "target")
    }
    connected_nodes = [
        node for node in filtered_nodes if str(node["id"]) in connected_node_ids
    ]
    filtered_graph = {"nodes": connected_nodes, "edges": filtered_edges}

    result = GLOSSARY_NETWORK(
        data=filtered_graph,
        key="glossary_network_graph",
        height=620,
        on_clicked_change=open_network_term_in_glossary,
    )


def parse_term_usage_summaries(json_text: str) -> TermUsageSummaries:
    """Parse the versioned precomputed-term-summary JSON artifact."""
    payload = json.loads(json_text)
    if not isinstance(payload, dict) or payload.get("format_version") != 1:
        raise ValueError("Precomputed term summaries must use format version 1.")

    records = payload.get("summaries")
    if not isinstance(records, dict):
        raise ValueError("Precomputed term summaries must contain a summaries object.")

    summaries: TermUsageSummaries = {}
    for term_key, record in records.items():
        if not isinstance(term_key, str) or not isinstance(record, dict):
            raise ValueError("Precomputed term summary entries must be keyed objects.")
        summary = record.get("summary")
        normalized_key = normalize_text(term_key).casefold()
        if not normalized_key or not isinstance(summary, str) or not summary.strip():
            raise ValueError("Precomputed term summary entries must contain nonempty text.")
        summaries[normalized_key] = summary.strip().replace(LEGACY_AR6_SOURCE_LABEL, AR6_SOURCE_LABEL)
    return summaries


@st.cache_data
def load_encrypted_assets(
    archive_path: str,
    modified_at_ns: int,
) -> tuple[str, Glossary, dict[str, list[dict[str, str | int]]], ExecutiveSummaries, str, TermUsageSummaries]:
    """Decrypt the report, glossary, network, executive summaries, prompt, and precomputed summaries."""
    del modified_at_ns
    path = Path(archive_path)
    if not path.is_file():
        raise FileNotFoundError("Report archive is unavailable.")

    try:
        decrypted_payload = get_fernet().decrypt(path.read_bytes())
        with ZipFile(BytesIO(decrypted_payload)) as archive:
            report_text = archive.read(REPORT_ARCHIVE_NAME).decode("utf-8")
            term_usage_summary_prompt = archive.read(TERM_USAGE_SUMMARY_PROMPT_ARCHIVE_NAME).decode("utf-8").replace(
                LEGACY_AR6_SOURCE_LABEL,
                AR6_SOURCE_LABEL,
            )
            term_usage_summaries = (
                parse_term_usage_summaries(archive.read(TERM_USAGE_SUMMARIES_ARCHIVE_NAME).decode("utf-8"))
                if TERM_USAGE_SUMMARIES_ARCHIVE_NAME in archive.namelist()
                else {}
            )
            network = parse_glossary_network(
                archive.read(GLOSSARY_NETWORK_ARCHIVE_MEMBER_NAME).decode("utf-8")
            )
            executive_summaries = {
                chapter: archive.read(archive_name).decode("utf-8")
                for chapter, archive_name in EXECUTIVE_SUMMARY_ARCHIVE_MEMBER_NAMES.items()
            }
            glossary: Glossary = {}
            for archive_name in GLOSSARY_ARCHIVE_MEMBER_NAMES:
                workbook = load_workbook(BytesIO(archive.read(archive_name)), read_only=True, data_only=True)
                worksheet = workbook.active
                headers = [normalize_text(cell.value).casefold() for cell in worksheet[1]]
                term_index = headers.index("term")
                explanation_index = headers.index("explanation")
                parent_index = headers.index("parent") if "parent" in headers else None
                source_index = headers.index("source")

                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    term = normalize_text(row[term_index] if term_index < len(row) else "")
                    explanation = normalize_text(row[explanation_index] if explanation_index < len(row) else "")
                    parent_value = row[parent_index] if parent_index is not None and parent_index < len(row) else None
                    parent = normalize_text(str(parent_value)) if parent_value is not None else ""
                    source = normalize_source_label(
                        normalize_text(row[source_index] if source_index < len(row) else "")
                    )
                    if not term or not explanation or not source:
                        continue
                    definition = (term, explanation, parent, source)
                    definitions = glossary.setdefault(term.casefold(), [])
                    if definition not in definitions:
                        definitions.append(definition)
                workbook.close()

            source_order = {"SRCities-SOD": 0, AR6_SOURCE_LABEL: 1}
            for definitions in glossary.values():
                definitions.sort(key=lambda definition: (source_order.get(definition[3], 2), definition[3]))

            return (
                report_text,
                glossary,
                network,
                executive_summaries,
                term_usage_summary_prompt,
                term_usage_summaries,
            )
    except (InvalidToken, RuntimeError, BadZipFile, KeyError, UnicodeDecodeError, ValueError) as error:
        raise RuntimeError("Report archive could not be loaded.") from error


@st.cache_data
def parse_markdown_sections(markdown_text: str) -> tuple[str, dict[str, list[str]]]:
    """Parse Markdown into major SPM sections for UI display."""
    lines = markdown_text.splitlines()

    title = "SRCities SPM"
    sections: dict[str, list[str]] = {key: [] for key in SECTION_LABELS}

    current_section: str | None = "Author Team"
    in_toc = False

    for raw_line in lines:
        text = raw_line.strip()
        if not text:
            continue

        normalized_text = normalize_text(text)
        lower_text = normalized_text.lower()

        if normalized_text.startswith("# "):
            title = normalized_text[2:].strip() or title
            continue

        if normalized_text.startswith("## "):
            heading = normalized_text[3:].strip().lower()
            in_toc = False
            if heading == "introduction":
                current_section = "Introduction"
            elif heading.startswith("section a"):
                current_section = "Section A"
            elif heading.startswith("section b"):
                current_section = "Section B"
            elif heading.startswith("section c"):
                current_section = "Section C"
            elif heading.startswith("section d"):
                current_section = "Section D"
            else:
                current_section = None
            continue

        if lower_text == "table of contents":
            in_toc = True
            continue

        if in_toc:
            continue

        if current_section is None:
            continue

        sections[current_section].append(normalized_text)

    return title, sections


def render_section(
    section_name: str,
    lines: list[str],
    glossary: Glossary,
) -> None:
    st.subheader(section_name)
    if not lines:
        st.info("No content parsed for this section.")
        return

    glossary_pattern = build_glossary_pattern(glossary)

    first_level_pattern = re.compile(r"^[A-D]\d\s")

    # Keep paragraph-level readability from the original DOCX.
    for idx, line in enumerate(lines):
        if first_level_pattern.match(line):
            st.markdown(f"###### **{line}**")
        else:
            rendered = linkify_glossary_terms(line, glossary, glossary_pattern)
            result = INLINE_TERM_TEXT(
                data={"html": f"<p>{rendered}</p>"},
                key=f"inline_terms_{section_name}_{idx}",
                on_clicked_change=lambda: None,
            )
            clicked_term = getattr(result, "clicked", "")
            if clicked_term:
                st.session_state.selected_term = normalize_text(str(clicked_term))


def render_executive_summary(chapter: str, markdown_text: str, glossary: Glossary) -> None:
    """Render executive summaries with the SPM heading scale and clickable glossary terms."""
    rendered_markdown = re.sub(
        r"^(#{1,4})\s",
        lambda match: f"{'#' * min(len(match.group(1)) + 2, 6)} ",
        markdown_text,
        flags=re.MULTILINE,
    )
    glossary_pattern = build_glossary_pattern(glossary)

    for index, block in enumerate(re.split(r"\n\s*\n", rendered_markdown)):
        if not block.strip():
            continue
        rendered_html = linkify_glossary_html(MARKDOWN_RENDERER.render(block), glossary, glossary_pattern)
        result = INLINE_TERM_TEXT(
            data={"html": rendered_html},
            key=f"inline_es_terms_{chapter.replace(' ', '_').lower()}_{index}",
            on_clicked_change=lambda: None,
        )
        clicked_term = getattr(result, "clicked", "")
        if clicked_term:
            st.session_state.selected_term = normalize_text(str(clicked_term))


def configured_access_passwords() -> tuple[str, ...]:
    """Read the approved password list from Community Cloud secrets or a local environment."""
    try:
        configured_passwords = st.secrets.get(USER_PASSWORDS_ENV_NAME, None)
    except FileNotFoundError:
        configured_passwords = None
    if configured_passwords is None:
        configured_passwords = os.environ.get(USER_PASSWORDS_ENV_NAME, "")

    if isinstance(configured_passwords, str):
        try:
            configured_passwords = ast.literal_eval(configured_passwords)
        except (SyntaxError, ValueError):
            return ()

    if not isinstance(configured_passwords, (list, tuple)):
        return ()
    return tuple(
        password
        for password in configured_passwords
        if isinstance(password, str) and password
    )


def client_ip_from_headers(headers: Mapping[str, object]) -> str:
    """Return the first valid client IP carried by trusted proxy request headers."""
    normalized_headers = {
        str(name).casefold(): str(value)
        for name, value in headers.items()
    }
    for header_name in ("x-forwarded-for", "cf-connecting-ip", "x-real-ip"):
        value = normalized_headers.get(header_name, "")
        for candidate in value.split(","):
            try:
                return str(ipaddress.ip_address(candidate.strip()))
            except ValueError:
                continue
    return "Unavailable"


def current_client_ip() -> str:
    """Get the client IP exposed by the current Streamlit request, if available."""
    try:
        return client_ip_from_headers(st.context.headers)
    except Exception:
        return "Unavailable"


def record_login_attempt(wrong_password: bool) -> None:
    """Write a non-blocking Notion audit record without ever passing the password."""
    try:
        from write2notion import write_login_attempt_to_notion

        write_login_attempt_to_notion(
            ip_address=current_client_ip(),
            wrong_password=wrong_password,
        )
    except Exception:
        LOGGER.warning("Login audit record could not be written.")


def validate_access_password() -> None:
    """Validate the password currently entered on the Settings page."""
    submitted_password = str(st.session_state.get("access_password_input", ""))
    expected_passwords = configured_access_passwords()
    password_is_valid = any(
        submitted_password and hmac.compare_digest(submitted_password, expected_password)
        for expected_password in expected_passwords
    )
    st.session_state.access_granted = password_is_valid
    record_login_attempt(wrong_password=not password_is_valid)
    st.session_state.access_password_input = ""


def render_access_settings() -> None:
    """Render the Settings page used to unlock the app."""
    st.divider()
    st.markdown("# Setting")

    expected_passwords = configured_access_passwords()
    if not expected_passwords:
        st.error("Access is currently unavailable. Please contact the app administrator.")
        return

    st.text_input(
        "Password",
        key="access_password_input",
        type="password",
        placeholder="Enter password to unlock",
    )
    st.button(
        "Validate Password",
        type="primary",
        key="validate_access_password",
        on_click=validate_access_password,
    )
    st.warning("Invalid or missing password. App functionality remains disabled.")


def render_user_guide() -> None:
    """Render the brief static guide for the checker."""
    if not (
        USER_GUIDE_PATH.is_file()
        and USER_GUIDE_SETTINGS_IMAGE_PATH.is_file()
        and USER_GUIDE_GLOSSARY_IMAGE_PATH.is_file()
        and USER_GUIDE_LLM_SUMMARY_IMAGE_PATH.is_file()
    ):
        st.error("The user guide is currently unavailable. Please contact the app administrator.")
        return
    guide_images = {
        "{{SETTING_CONTROLS_IMAGE}}": USER_GUIDE_SETTINGS_IMAGE_PATH,
        "{{GLOSSARY_OVERVIEW_IMAGE}}": USER_GUIDE_GLOSSARY_IMAGE_PATH,
        "{{LLM_SUMMARY_IMAGE}}": USER_GUIDE_LLM_SUMMARY_IMAGE_PATH,
    }
    guide_html = USER_GUIDE_PATH.read_text(encoding="utf-8")
    for marker, image_path in guide_images.items():
        image_data = base64.b64encode(image_path.read_bytes()).decode("ascii")
        guide_html = guide_html.replace(marker, f"data:image/png;base64,{image_data}")
    st.html(guide_html)


def main() -> None:
    st.set_page_config(page_title="SRCities SPM Viewer", layout="wide")
    if not st.session_state.get("access_granted", False):
        st.sidebar.markdown("<h1 style='color: #0076A8;'>SPM and ESs checker</h1>", unsafe_allow_html=True)
        st.sidebar.markdown("**Internal Use by xWG TSU**")
        st.sidebar.radio(
            "Settings",
            [SETTINGS_ACCESS_LABEL],
            index=0,
            key="settings_nav_section",
        )
        render_access_settings()
        st.stop()

    st.html(
        """
        <style>
        [data-testid="stAppViewContainer"] {
            background:
                linear-gradient(135deg, rgba(0, 118, 168, 0.08), transparent 38%),
                linear-gradient(180deg, #f8fcfe 0%, #eef7fb 100%);
        }
        .st-key-top_panel,
        .st-key-bottom_panel {
            border: 1px solid #b9d8e7;
            border-radius: 6px;
            box-shadow: 0 8px 22px rgba(21, 61, 82, 0.08);
        }
        .st-key-top_panel {
            background: rgba(255, 255, 255, 0.92);
            border-top: 3px solid #0076a8;
        }
        .st-key-bottom_panel {
            background: rgba(241, 248, 252, 0.94);
            border-top: 3px solid #3a98c1;
        }
        </style>
        """
    )

    try:
        archive_modified_at_ns = ENCRYPTED_REPORT_PATH.stat().st_mtime_ns
        (
            markdown_text,
            glossary,
            glossary_network,
            executive_summaries,
            _,
            term_usage_summaries,
        ) = load_encrypted_assets(
            str(ENCRYPTED_REPORT_PATH), archive_modified_at_ns
        )
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        st.error("The report is currently unavailable. Please contact the app administrator.")
        st.stop()

    doc_title, sections = parse_markdown_sections(markdown_text)

    if "selected_term" not in st.session_state:
        st.session_state.selected_term = ""
    if "active_navigation_group" not in st.session_state:
        st.session_state.active_navigation_group = "spm"

    st.sidebar.markdown(
        """
        <style>
        [data-testid="stSidebarUserContent"] > div[data-testid="stVerticalBlock"] {
            gap: 0.35rem;
        }
        [data-testid="stSidebarUserContent"] h1 {
            margin-bottom: 0.15rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.sidebar.markdown("<h1 style='color: #0076A8;'>SPM and ESs checker</h1>", unsafe_allow_html=True)
    st.sidebar.markdown("**Internal Use by xWG TSU**")
    user_guide_choice = st.sidebar.radio(
        "User Guide",
        [USER_GUIDE_LABEL],
        index=0 if st.session_state.active_navigation_group == "user_guide" else None,
        key="user_guide_section",
        on_change=select_navigation_group,
        args=("user_guide",),
        label_visibility="collapsed",
    )
    executive_summary_choice = st.sidebar.radio(
        "Executive summaries",
        list(EXECUTIVE_SUMMARY_ARCHIVE_MEMBER_NAMES),
        index=0 if st.session_state.active_navigation_group == "executive_summaries" else None,
        key="executive_summary_section",
        on_change=select_navigation_group,
        args=("executive_summaries",),
    )
    spm_choice = st.sidebar.radio(
        "SPM",
        SPM_SECTION_LABELS,
        index=0 if st.session_state.active_navigation_group == "spm" else None,
        key="spm_nav_section",
        on_change=select_navigation_group,
        args=("spm",),
    )
    glossary_choice = st.sidebar.radio(
        "Glossary",
        GLOSSARY_SECTION_LABELS,
        index=0 if st.session_state.active_navigation_group == "glossary" else None,
        key="glossary_nav_section",
        on_change=select_navigation_group,
        args=("glossary",),
    )
    if st.session_state.active_navigation_group == "user_guide":
        choice = user_guide_choice
    elif st.session_state.active_navigation_group == "executive_summaries":
        choice = executive_summary_choice
    elif st.session_state.active_navigation_group == "spm":
        choice = spm_choice
    elif st.session_state.active_navigation_group == "glossary":
        choice = glossary_choice
    else:
        choice = spm_choice
    showing_executive_summary = (
        st.session_state.active_navigation_group == "executive_summaries"
        and choice in executive_summaries
    )
    if choice == USER_GUIDE_LABEL:
        render_user_guide()
        return

    split_ratio = st.sidebar.slider("Divider position (%)", min_value=20, max_value=80, value=60, key="split_ratio")
    panel_height = st.sidebar.slider("Panel height", min_value=420, max_value=1000, value=640, step=20, key="panel_height")

    if choice == GLOSSARY_NETWORK_LABEL:
        render_glossary_network(glossary_network)
        return
    if choice == SANKEY_DIAGRAM_LABEL:
        render_sankey_diagram(sections)
        return

    left_col, right_col = st.columns([split_ratio, 100 - split_ratio], gap="small")
    with left_col:
        # Fixed-height container provides its own vertical scrollbar.
        top_panel = st.container(border=True, key="top_panel", height=panel_height)
        with top_panel:
            if showing_executive_summary:
                render_executive_summary(choice, executive_summaries[choice], glossary)
            elif choice == GLOSSARY_OVERVIEW_LABEL:
                render_glossary(glossary, sections, executive_summaries, term_usage_summaries)
            else:
                current_lines = sections.get(choice, [])
                render_section(choice, current_lines, glossary)
    with right_col:
        # Independent fixed-height container for definition/usage scrolling.
        bottom_panel = st.container(border=True, key="bottom_panel", height=panel_height)
        with bottom_panel:
            if showing_executive_summary or choice == GLOSSARY_OVERVIEW_LABEL:
                render_term_details(
                    get_selected_term(),
                    "",
                    sections,
                    glossary,
                    executive_summaries,
                    term_usage_summaries,
                    include_current_section=True,
                )
            else:
                render_term_details(
                    get_selected_term(),
                    choice,
                    sections,
                    glossary,
                    executive_summaries,
                    term_usage_summaries,
                    include_current_section=True,
                )


if __name__ == "__main__":
    main()
