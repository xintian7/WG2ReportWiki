#!/usr/bin/env python3
"""Merge the AR6 working-group glossaries and AR7 SOD glossary into Excel.

The source ``.doc`` files are UTF-8 HTML exports. Terms carrying a marker such
as ``{↑ Scenario}`` are linked to that parent; child terms are derived from the
inverse relationship. Duplicate terms are merged case-insensitively, with AR7
SOD explanations taking precedence over AR6 explanations. Appended note sections
are omitted from AR6 explanations.

Example:
    /opt/anaconda3/envs/tsu_repwiki/bin/python script/merge_ar6_ar7sod_glossaries.py
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass, field
from html.parser import HTMLParser
from itertools import product
import json
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_ORDER = ("AR6WGI", "AR6WGII", "AR6WGIII", "AR7SOD")
DEFAULT_WGI_PATH = REPO_ROOT / "data/Glossary/AR6WGI_glossary.doc"
DEFAULT_WGII_PATH = REPO_ROOT / "data/Glossary/AR6WGII_glossary.doc"
DEFAULT_WGIII_PATH = REPO_ROOT / "data/Glossary/AR6WGIII_glossary.doc"
DEFAULT_AR7SOD_PATH = REPO_ROOT / "data/Glossary/AR7SOD_Glossary.xlsx"
DEFAULT_SOD_JSON_PATH = REPO_ROOT / "data/srsod-inspection.json"
DEFAULT_OUTPUT_PATH = REPO_ROOT / "data/Glossary/AR6_AR7SOD_Glossary.xlsx"
ENTRY_STYLE_RE = re.compile(
    r"(?:^|;)\s*margin-bottom\s*:\s*15px(?:;|$)",
    re.IGNORECASE,
)
PARENT_MARKER_RE = re.compile(r"\s*\{\s*↑\s*(?P<parent>.*?)\s*\}\s*$")
NOTE_START_RE = re.compile(r"(?:^|\s)\[?Note(?:\s+\d+)?\s*:", re.IGNORECASE)
PARENTHETICAL_RE = re.compile(r"\(([^()]*)\)")
BRACKETED_ALTERNATIVES_RE = re.compile(r"^(?P<head>.+?)\s*\[(?P<options>[^][]+)\]\s*$")
OR_LIST_SEPARATOR_RE = re.compile(r"\s*,\s*(?:or\s+)?|\s+or\s+", re.IGNORECASE)
PLAIN_PHRASE_RE = re.compile(r"^[^\W\d_]+(?:[-'][^\W\d_]+)*(?: [^\W\d_]+(?:[-'][^\W\d_]+)*)*$")
SPELLING_PAIRS = (
    ("behavior", "behaviour"),
    ("center", "centre"),
    ("fertilization", "fertilisation"),
    ("modeling", "modelling"),
    ("organization", "organisation"),
    ("urbanization", "urbanisation"),
)
INLINE_ALIAS_RE = re.compile(
    r"^\s*,?\s*(?:(?:is|are|was|were|can be)\s+)?"
    r"(?:(?:also|often|sometimes|typically|commonly|most often)\s+)?"
    r"(?:known as|called|referred to as)\s+"
    r"(?P<alias>[^;:.!?()]{1,120})",
    re.IGNORECASE,
)
PARENTHETICAL_ALIAS_RE = re.compile(
    r"^(?:(?:also|formerly|commonly|sometimes|typically)\s+)?"
    r"(?:known as|called|referred to as)\s+(?P<alias>.+)$",
    re.IGNORECASE,
)
APPOSITIVE_ALIAS_RE = re.compile(
    r"^\s*,\s*(?:or|alternatively)\s+(?P<alias>[^,;:.!?()]{1,80})(?=[,;:.!?])",
    re.IGNORECASE,
)
ALIAS_CONTEXT_RE = re.compile(
    r"\s+(?:in|when|where|which|that|to describe|for example)\b.*$",
    re.IGNORECASE,
)
QUOTED_ALIAS_RE = re.compile(r"[\"“‘]([^\"”’]+)[\"”’]")
ADJECTIVAL_COMPOUND_MODIFIERS = {
    "economic",
    "environmental",
    "global",
    "human",
    "indigenous",
    "local",
    "national",
    "political",
    "regional",
    "rural",
    "social",
    "urban",
}
ACRONYM_STOPWORDS = {"a", "an", "and", "for", "in", "of", "or", "the", "to"}


@dataclass(frozen=True)
class SourceGlossaryEntry:
    """One term definition and optional parent from one working group."""

    term: str
    explanation: str
    parent: str
    source: str


@dataclass
class MergedGlossaryEntry:
    """Definitions and relationships collected for one case-insensitive term."""

    term: str
    sources: set[str] = field(default_factory=set)
    explanations: dict[str, list[str]] = field(default_factory=dict)
    parents: dict[str, str] = field(default_factory=dict)

    def add(self, entry: SourceGlossaryEntry) -> None:
        self.sources.add(entry.source)
        source_explanations = self.explanations.setdefault(entry.source, [])
        if entry.explanation and entry.explanation not in source_explanations:
            source_explanations.append(entry.explanation)
        if entry.parent:
            self.parents.setdefault(entry.parent.casefold(), entry.parent)


class GlossaryDocumentParser(HTMLParser):
    """Parse glossary entry containers from an IPCC HTML document export."""

    def __init__(self, source: str) -> None:
        super().__init__(convert_charrefs=True)
        self.source = source
        self.entries: list[SourceGlossaryEntry] = []
        self._entry_depth = 0
        self._section_index = -1
        self._parts: list[list[str]] = [[], []]

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if self._entry_depth == 0:
            if tag == "div" and ENTRY_STYLE_RE.search(attributes.get("style") or ""):
                self._entry_depth = 1
                self._section_index = -1
                self._parts = [[], []]
            return

        if tag == "div":
            if self._entry_depth == 1:
                self._section_index += 1
            self._entry_depth += 1
        elif tag in {"br", "li", "p"} and self._section_index == 1:
            self._parts[1].append(" ")

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        del attrs
        if self._entry_depth and tag in {"br", "li", "p"} and self._section_index == 1:
            self._parts[1].append(" ")

    def handle_endtag(self, tag: str) -> None:
        if self._entry_depth == 0 or tag != "div":
            return
        self._entry_depth -= 1
        if self._entry_depth == 0:
            self._finish_entry()

    def handle_data(self, data: str) -> None:
        if self._entry_depth and self._section_index in (0, 1):
            self._parts[self._section_index].append(data)

    def _finish_entry(self) -> None:
        heading = normalize_text("".join(self._parts[0]))
        explanation = without_notes("".join(self._parts[1]))
        parent_match = PARENT_MARKER_RE.search(heading)
        parent = normalize_text(parent_match.group("parent")) if parent_match else ""
        term = normalize_text(heading[: parent_match.start()] if parent_match else heading)
        if not term:
            raise ValueError(f"Encountered a glossary entry without a term in {self.source}.")
        self.entries.append(SourceGlossaryEntry(term, explanation, parent, self.source))


def normalize_text(value: object) -> str:
    """Convert source content to normalized plain text."""
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def without_notes(value: object) -> str:
    """Remove appended Note/Note N sections from a glossary explanation."""
    explanation = normalize_text(value)
    note_match = NOTE_START_RE.search(explanation)
    return explanation[: note_match.start()].rstrip() if note_match else explanation


def unique_equivalents(values: list[str]) -> list[str]:
    """Return normalized equivalent terms without case-insensitive duplicates."""
    equivalents: list[str] = []
    seen: set[str] = set()
    for value in values:
        equivalent = normalize_text(value).strip(" ,")
        key = equivalent.casefold()
        if equivalent and key not in seen:
            seen.add(key)
            equivalents.append(equivalent)
    return equivalents


def expand_bracketed_alternatives(term: str) -> list[str]:
    """Combine a parenthetical synonym with each bracketed noun alternative."""
    match = BRACKETED_ALTERNATIVES_RE.fullmatch(term)
    if match is None:
        return []

    head = normalize_text(match.group("head"))
    head_match = PARENTHETICAL_RE.search(head)
    if head_match is None:
        heads = [head]
    else:
        primary = normalize_text(f"{head[:head_match.start()]} {head[head_match.end():]}")
        synonym = normalize_text(head_match.group(1))
        heads = [primary, synonym]

    nouns = [
        "emission" if option.casefold() == "emissions" else option
        for option in split_or_list(match.group("options"))
    ]
    return unique_equivalents([f"{head_variant} {noun}" for head_variant in heads for noun in nouns])


def split_top_level_slashes(term: str) -> list[str]:
    """Expand slash alternatives while leaving slashes inside parentheses intact."""
    parts: list[str] = []
    depth = 0
    part_start = 0
    for index, character in enumerate(term):
        if character == "(":
            depth += 1
        elif character == ")":
            depth = max(0, depth - 1)
        elif character == "/" and depth == 0:
            prefix = term[:index].rstrip().casefold()
            suffix = term[index + 1 :].lstrip().casefold()
            if prefix.endswith("and") and suffix.startswith("or "):
                continue
            parts.append(term[part_start:index].strip())
            part_start = index + 1
    parts.append(term[part_start:].strip())
    if len(parts) == 1:
        return parts

    if len(parts) != 2 or not parts[1][:1].islower():
        return unique_equivalents(parts)

    left_words = normalize_text(PARENTHETICAL_RE.sub("", parts[0])).split()
    right_words = normalize_text(PARENTHETICAL_RE.sub("", parts[1])).split()
    if not left_words or not right_words:
        return unique_equivalents(parts)

    left_root = left_words[0].casefold().rstrip("s")
    right_root = right_words[0].casefold().rstrip("s")
    if left_root == right_root:
        return unique_equivalents(parts)

    if len(left_words) == 1 and len(right_words) > 1:
        parts[0] = f"{parts[0]} {' '.join(right_words[1:])}"
    elif len(left_words) > 1:
        parts[1] = f"{' '.join(left_words[:-1])} {parts[1]}"
        if len(right_words) > 1:
            parts[0] = f"{parts[0]} {' '.join(right_words[1:])}"

    trailing_parentheticals = " ".join(
        match.group(0) for match in PARENTHETICAL_RE.finditer(parts[1])
    )
    if trailing_parentheticals and "(" not in parts[0]:
        parts[0] = f"{parts[0]} {trailing_parentheticals}"
    return unique_equivalents(parts)


def split_or_list(text: str) -> list[str]:
    """Split comma-and-or alternatives while preserving their source order."""
    return [part.strip() for part in OR_LIST_SEPARATOR_RE.split(text) if part.strip()]


def parenthetical_options(content: str) -> list[str]:
    """Expand alternatives within one optional parenthetical qualifier."""
    if not content.casefold().startswith("of "):
        return [content]

    options = []
    for item in split_or_list(content[3:]):
        if item.casefold() == "ice sheets":
            item = "ice sheet"
        options.append(f"of {item}")
    return options


def is_abbreviation(content: str) -> bool:
    """Return whether terminal parenthetical text is acronym-like."""
    if re.fullmatch(r"[A-Z](?:-[A-Z])+(?:\s+events?)?", content):
        return True
    if " " in content:
        return False

    letters = [character for character in content if character.isalpha()]
    uppercase_count = sum(character.isupper() for character in letters)
    return bool(letters) and (
        uppercase_count >= 2
        or any(character.isdigit() for character in content)
        or (uppercase_count >= 1 and len(letters) <= 4)
        or (any(ord(character) > 127 for character in letters) and len(letters) <= 4)
    )


def expand_parentheticals(term: str) -> list[str]:
    """Expand parenthetical acronyms, aliases, and optional qualifier combinations."""
    matches = list(PARENTHETICAL_RE.finditer(term))
    if not matches:
        return []

    for match in reversed(matches):
        content = normalize_text(match.group(1))
        if not is_abbreviation(content):
            continue
        base_term = normalize_text(f"{term[:match.start()]} {term[match.end():]}")
        base_equivalents = expand_parentheticals(base_term) or [base_term]
        suffix = normalize_text(term[match.end() :])
        abbreviation_head = content
        expanded_head = normalize_text(term[: match.start()])
        if content.casefold() == "co2" and expanded_head.casefold().endswith("carbon dioxide"):
            leading_modifier = expanded_head[: -len("carbon dioxide")].rstrip()
            abbreviation_head = normalize_text(f"{leading_modifier} {content}")
        abbreviation = normalize_text(f"{abbreviation_head} {suffix}")
        return unique_equivalents([*base_equivalents, abbreviation])

    for match in matches:
        content = normalize_text(match.group(1))
        if not content.casefold().startswith("or "):
            continue
        base_term = normalize_text(f"{term[:match.start()]} {term[match.end():]}")
        prefix = term[: match.start()].rstrip()
        replacement_prefix = re.sub(r"\S+$", content[3:].strip(), prefix)
        replacement_term = normalize_text(f"{replacement_prefix} {term[match.end():]}")
        equivalents = []
        for variant in (base_term, replacement_term):
            equivalents.extend(expand_parentheticals(variant) or [variant])
        return unique_equivalents(equivalents)

    if len(matches) == 1 and matches[0].end() == len(term):
        content = normalize_text(matches[0].group(1))
        base_term = normalize_text(term[: matches[0].start()])
        if content.casefold().startswith(("of the ", "in relation to ")):
            return unique_equivalents([base_term, f"{base_term} {content}"])
        if content.casefold().startswith("also "):
            aliases = split_or_list(content[5:])
            return unique_equivalents([base_term, *aliases])
        if "/" in content:
            aliases = [part.strip() for part in content.split("/")]
            return unique_equivalents([base_term, *aliases])
        if " or " in content.casefold() and not content.casefold().startswith("of "):
            return unique_equivalents([base_term, *split_or_list(content)])

    segments: list[str] = []
    option_groups: list[list[str]] = []
    cursor = 0
    for match in matches:
        segments.append(term[cursor : match.start()])
        option_groups.append(parenthetical_options(normalize_text(match.group(1))))
        cursor = match.end()
    segments.append(term[cursor:])

    equivalents: list[str] = []
    for mask in range(1, 1 << len(option_groups)):
        selected_groups = [
            options if mask & (1 << index) else [""]
            for index, options in enumerate(option_groups)
        ]
        for selections in product(*selected_groups):
            pieces = []
            for segment, selection in zip(segments, selections):
                pieces.extend((segment, selection))
            pieces.append(segments[-1])
            equivalent = normalize_text("".join(pieces))
            if term.lstrip().startswith("(") and equivalent:
                equivalent = equivalent[:1].lower() + equivalent[1:]
            equivalents.append(equivalent)
    return unique_equivalents(equivalents)


def equivalent_terms(term: str) -> list[str]:
    """Derive searchable equivalents from parenthetical and slash notation."""
    bracketed_equivalents = expand_bracketed_alternatives(term)
    if bracketed_equivalents:
        return bracketed_equivalents

    slash_variants = split_top_level_slashes(term)
    if len(slash_variants) == 1:
        return expand_parentheticals(term)

    equivalents: list[str] = []
    for variant in slash_variants:
        equivalents.extend(expand_parentheticals(variant) or [variant])
    return unique_equivalents(equivalents)


def pluralize_word(word: str) -> str:
    """Return a conservative English plural for a glossary word."""
    word_key = word.casefold()
    irregular = {"city": "cities", "person": "people"}
    if word_key in irregular:
        plural = irregular[word_key]
    elif word_key.endswith(("s", "x", "z", "ch", "sh")):
        plural = f"{word}es"
    elif len(word) > 1 and word_key.endswith("y") and word_key[-2] not in "aeiou":
        plural = f"{word[:-1]}ies"
    else:
        plural = f"{word}s"
    return plural.capitalize() if word[:1].isupper() else plural


def singularize_word(word: str) -> str:
    """Return a conservative English singular for a glossary word."""
    word_key = word.casefold()
    irregular = {"cities": "city", "people": "person"}
    if word_key in irregular:
        singular = irregular[word_key]
    elif len(word) > 3 and word_key.endswith("ies"):
        singular = f"{word[:-3]}y"
    elif word_key.endswith(("ses", "xes", "zes", "ches", "shes")):
        singular = word[:-2]
    elif len(word) > 2 and word_key.endswith("s") and not word_key.endswith(("ss", "us", "is")):
        singular = word[:-1]
    else:
        singular = word
    return singular.capitalize() if word[:1].isupper() else singular


def direct_linguistic_variants(term: str) -> list[str]:
    """Generate number, compound, hyphenation, and spelling variants."""
    if not PLAIN_PHRASE_RE.fullmatch(term):
        return []

    words = term.split()
    variants = [
        " ".join([*words[:-1], singularize_word(words[-1])]),
        " ".join([*words[:-1], pluralize_word(words[-1])]),
    ]
    if "-" in term:
        variants.append(term.replace("-", " "))
    elif len(words) == 2:
        variants.append("-".join(words))

    if len(words) == 2 and words[0].casefold() not in ADJECTIVAL_COMPOUND_MODIFIERS:
        modifier, head = words
        moved_modifier = pluralize_word(singularize_word(modifier))
        if modifier.istitle():
            moved_modifier = moved_modifier.casefold()
        variants.append(f"{head} of {moved_modifier}")
    elif len(words) == 3 and words[1].casefold() == "of":
        head, _, modifier = words
        variants.append(f"{singularize_word(modifier)} {head}")

    for index, word in enumerate(words):
        for american, british in SPELLING_PAIRS:
            replacements = {american: british, british: american}
            replacement = replacements.get(word.casefold())
            if replacement:
                spelling_variant = words.copy()
                spelling_variant[index] = (
                    replacement.capitalize() if word[:1].isupper() else replacement
                )
                variants.append(" ".join(spelling_variant))
    return unique_equivalents(variants)


def linguistic_candidates(term: str) -> list[str]:
    """Generate two rounds of deterministic variants from a term and column F."""
    seeds = unique_equivalents([term, *equivalent_terms(term)])
    candidates: list[str] = []
    frontier = seeds
    for _ in range(2):
        next_frontier: list[str] = []
        for variant in frontier:
            next_frontier.extend(direct_linguistic_variants(variant))
        next_frontier = unique_equivalents(next_frontier)
        candidates.extend(next_frontier)
        frontier = next_frontier

    excluded = {seed.casefold() for seed in seeds}
    return [
        candidate
        for candidate in unique_equivalents(candidates)
        if candidate.casefold() not in excluded
    ]


def load_sod_text_blocks(input_path: Path) -> list[str]:
    """Load eligible paragraph and figure text from the full SOD inspection JSON."""
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    reports = payload.get("reports") if isinstance(payload, dict) else None
    if not isinstance(reports, list):
        raise ValueError("SOD inspection JSON must contain a reports list.")

    blocks: list[str] = []

    def collect(node: dict[str, object]) -> None:
        title = node.get("title")
        normalized_title = normalize_text(title).casefold() if isinstance(title, str) else ""
        if normalized_title == "references" or normalized_title.startswith("supplementary material"):
            return

        kind = node.get("kind")
        value = (
            node.get("text")
            if kind == "paragraph"
            else node.get("explanation")
            if kind == "figure"
            else None
        )
        if isinstance(value, str) and normalize_text(value):
            blocks.append(normalize_text(value))
        children = node.get("children", [])
        if not isinstance(children, list):
            raise ValueError("A SOD report node has non-list children.")
        for child in children:
            if isinstance(child, dict):
                collect(child)

    for report in reports:
        if not isinstance(report, dict) or not isinstance(report.get("tree"), dict):
            raise ValueError("Each SOD inspection report must contain a tree object.")
        collect(report["tree"])
    if not blocks:
        raise RuntimeError(f"No eligible SOD text was parsed from {input_path}.")
    return blocks


def corpus_confirmed_linguistic_variants(
    glossary: dict[str, MergedGlossaryEntry],
    text_blocks: list[str],
) -> dict[str, list[str]]:
    """Retain deterministic alternatives that occur in the full SOD text."""
    candidates_by_term = {
        term_key: linguistic_candidates(entry.term)
        for term_key, entry in glossary.items()
    }
    candidate_lookup = {
        candidate.casefold(): candidate
        for candidates in candidates_by_term.values()
        for candidate in candidates
    }
    found_keys: set[str] = set()
    if candidate_lookup:
        pattern = re.compile(
            r"(?<![\w])(?:"
            + "|".join(
                re.escape(candidate)
                for candidate in sorted(candidate_lookup, key=len, reverse=True)
            )
            + r")(?![\w])",
            re.IGNORECASE,
        )
        for block in text_blocks:
            found_keys.update(match.group(0).casefold() for match in pattern.finditer(block))

    glossary_keys = set(glossary)
    confirmed: dict[str, list[str]] = {}
    for term_key, candidates in candidates_by_term.items():
        confirmed[term_key] = [
            candidate
            for candidate in candidates
            if candidate.casefold() in found_keys
            and (candidate.casefold() not in glossary_keys or candidate.casefold() == term_key)
        ]
    return confirmed


def normalize_aliases(value: str) -> list[str]:
    """Normalize one explicit alias expression into short candidate phrases."""
    quoted_aliases = QUOTED_ALIAS_RE.findall(value)
    if quoted_aliases:
        return unique_equivalents(quoted_aliases)

    alias = normalize_text(value).strip(" \"'“”‘’,-")
    alias = ALIAS_CONTEXT_RE.sub("", alias).strip(" \"'“”‘’,-")
    if not alias:
        return []

    alternatives = split_or_list(alias)
    if len(alternatives) > 1 and all(is_abbreviation(item) for item in alternatives):
        aliases = alternatives
    else:
        aliases = [alias]
    return [
        candidate
        for candidate in unique_equivalents(aliases)
        if len(candidate.split()) <= 8 and any(character.isalpha() for character in candidate)
    ]


def abbreviation_matches_phrase(abbreviation: str, phrase: str) -> bool:
    """Return whether an unmarked parenthetical abbreviation fits its phrase."""
    abbreviation_key = re.sub(r"[^A-Za-z0-9]", "", abbreviation).casefold()
    phrase_words = [
        word
        for word in re.findall(r"[^\W_]+", phrase)
        if word.casefold() not in ACRONYM_STOPWORDS
    ]
    initials = "".join(word[0] for word in phrase_words).casefold()
    return bool(initials) and abbreviation_key in {initials, f"{initials}s"}


def standalone_alias_subject(block: str, term_start: int) -> bool:
    """Return whether a matched term is the complete subject of an alias clause."""
    prefix = block[:term_start]
    boundary = max(prefix.rfind(mark) for mark in ".!?;:")
    fragment = normalize_text(prefix[boundary + 1 :]).casefold()
    return fragment in {"", "a", "an", "the", "this", "these", "those", "such"}


def aliases_after_term(
    text_after_term: str,
    matched_term: str,
    allow_inline_alias: bool,
) -> list[str]:
    """Extract aliases introduced immediately after a term occurrence."""
    parenthetical_match = re.match(r"^\s*\(([^()]{1,120})\)", text_after_term)
    if parenthetical_match:
        content = normalize_text(parenthetical_match.group(1))
        signal_match = PARENTHETICAL_ALIAS_RE.fullmatch(content)
        if signal_match:
            return normalize_aliases(signal_match.group("alias"))
        if is_abbreviation(content) and abbreviation_matches_phrase(content, matched_term):
            return [content]

    if allow_inline_alias:
        inline_match = INLINE_ALIAS_RE.match(text_after_term)
        if inline_match:
            return normalize_aliases(inline_match.group("alias"))

    appositive_match = APPOSITIVE_ALIAS_RE.match(text_after_term)
    if appositive_match:
        aliases = normalize_aliases(appositive_match.group("alias"))
        return [alias for alias in aliases if is_abbreviation(alias)]
    return []


def explicit_sod_aliases(
    glossary: dict[str, MergedGlossaryEntry],
    text_blocks: list[str],
) -> dict[str, list[str]]:
    """Mine aliases attached to glossary terms by explicit SOD wording."""
    term_keys_by_seed: dict[str, set[str]] = {}
    excluded_by_term: dict[str, set[str]] = {}
    for term_key, entry in glossary.items():
        seeds = unique_equivalents([entry.term, *equivalent_terms(entry.term)])
        excluded_by_term[term_key] = {seed.casefold() for seed in seeds}
        for seed in seeds:
            if len(seed) > 1:
                term_keys_by_seed.setdefault(seed.casefold(), set()).add(term_key)

    aliases: dict[str, list[str]] = {term_key: [] for term_key in glossary}
    if not term_keys_by_seed:
        return aliases
    term_pattern = re.compile(
        r"(?<![\w])(?:"
        + "|".join(re.escape(seed) for seed in sorted(term_keys_by_seed, key=len, reverse=True))
        + r")(?![\w])",
        re.IGNORECASE,
    )
    glossary_keys = set(glossary)
    for block in text_blocks:
        for match in term_pattern.finditer(block):
            matched_seed = match.group(0).casefold()
            for term_key in term_keys_by_seed[matched_seed]:
                for alias in aliases_after_term(
                    block[match.end() : match.end() + 180],
                    match.group(0),
                    standalone_alias_subject(block, match.start()),
                ):
                    alias_key = alias.casefold()
                    if alias_key in excluded_by_term[term_key] or alias_key in glossary_keys:
                        continue
                    aliases[term_key].append(alias)

    return {
        term_key: unique_equivalents(term_aliases)
        for term_key, term_aliases in aliases.items()
    }


def scan_equivalent_terms_2(
    glossary: dict[str, MergedGlossaryEntry],
    text_blocks: list[str],
) -> dict[str, list[str]]:
    """Combine corpus-confirmed linguistic variants and explicit SOD aliases."""
    linguistic = corpus_confirmed_linguistic_variants(glossary, text_blocks)
    explicit = explicit_sod_aliases(glossary, text_blocks)
    return {
        term_key: unique_equivalents([*linguistic[term_key], *explicit[term_key]])
        for term_key in glossary
    }


def parse_glossary(input_path: Path, source: str) -> list[SourceGlossaryEntry]:
    """Parse all glossary entries from one UTF-8 HTML document export."""
    parser = GlossaryDocumentParser(source)
    parser.feed(input_path.read_text(encoding="utf-8"))
    parser.close()
    if not parser.entries:
        raise RuntimeError(f"No glossary entries were parsed from {input_path}.")
    return parser.entries


def parse_glossary_workbook(
    input_path: Path, source: str
) -> list[SourceGlossaryEntry]:
    """Parse term, explanation, and parent columns from a glossary workbook."""
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = {
            normalize_text(cell.value).casefold(): index
            for index, cell in enumerate(worksheet[1])
            if normalize_text(cell.value)
        }
        missing_headers = {"term", "explanation"} - headers.keys()
        if missing_headers:
            missing = ", ".join(sorted(missing_headers))
            raise ValueError(f"Missing required column(s) in {input_path}: {missing}")

        parent_index = headers.get("parent")
        entries = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            term = normalize_text(row[headers["term"]])
            if not term:
                continue
            explanation = normalize_text(row[headers["explanation"]])
            parent = (
                normalize_text(row[parent_index])
                if parent_index is not None and parent_index < len(row)
                else ""
            )
            entries.append(SourceGlossaryEntry(term, explanation, parent, source))
    finally:
        workbook.close()

    if not entries:
        raise RuntimeError(f"No glossary entries were parsed from {input_path}.")
    return entries


def merge_glossaries(
    entries_by_source: dict[str, list[SourceGlossaryEntry]],
) -> dict[str, MergedGlossaryEntry]:
    """Merge case-insensitive terms while retaining source-specific definitions."""
    merged: dict[str, MergedGlossaryEntry] = {}
    for source in SOURCE_ORDER:
        for entry in entries_by_source[source]:
            term_key = entry.term.casefold()
            merged_entry = merged.setdefault(term_key, MergedGlossaryEntry(term=entry.term))
            merged_entry.add(entry)
    return merged


def source_label(entry: MergedGlossaryEntry) -> str:
    """Format the working groups in which a term is defined."""
    return "+".join(source for source in SOURCE_ORDER if source in entry.sources)


def format_explanation(entry: MergedGlossaryEntry) -> str:
    """Prefer AR7 SOD; otherwise retain distinct labelled AR6 definitions."""
    if "AR7SOD" in entry.sources:
        return "\n\n".join(entry.explanations.get("AR7SOD", []))

    explanation_sources: dict[str, list[str]] = {}
    for source in SOURCE_ORDER[:-1]:
        for explanation in entry.explanations.get(source, []):
            explanation_sources.setdefault(explanation, []).append(source)

    if not explanation_sources:
        return ""
    if len(explanation_sources) == 1:
        return next(iter(explanation_sources))
    return "\n\n".join(
        f"{'+'.join(sources)}:\n{explanation}"
        for explanation, sources in explanation_sources.items()
    )


def build_child_terms(
    glossary: dict[str, MergedGlossaryEntry],
) -> tuple[dict[str, dict[str, str]], list[tuple[str, str]]]:
    """Invert parent links into child lists and report unresolved parents."""
    children: dict[str, dict[str, str]] = {term_key: {} for term_key in glossary}
    unresolved: list[tuple[str, str]] = []
    for child_key, entry in glossary.items():
        for parent_key, parent in entry.parents.items():
            if parent_key not in glossary:
                unresolved.append((entry.term, parent))
                continue
            children[parent_key].setdefault(child_key, entry.term)
    return children, unresolved


def workbook_rows(
    glossary: dict[str, MergedGlossaryEntry],
    equivalent_terms_2: dict[str, list[str]] | None = None,
) -> tuple[list[tuple[str, str, str, str, str, str, str]], list[tuple[str, str]]]:
    """Build alphabetized rows and derive inverse child relationships."""
    child_terms, unresolved = build_child_terms(glossary)
    equivalent_terms_2 = equivalent_terms_2 or {}
    rows = []
    for term_key in sorted(glossary):
        entry = glossary[term_key]
        rows.append(
            (
                entry.term,
                format_explanation(entry),
                source_label(entry),
                "\n".join(sorted(entry.parents.values(), key=str.casefold)),
                "\n".join(sorted(child_terms[term_key].values(), key=str.casefold)),
                "\n".join(equivalent_terms(entry.term)),
                "\n".join(equivalent_terms_2.get(term_key, [])),
            )
        )
    return rows, unresolved


def write_workbook(
    rows: list[tuple[str, str, str, str, str, str, str]],
    output_path: Path,
) -> None:
    """Write merged glossary definitions and relationships to Excel."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Glossary"
    worksheet.append(
        [
            "Term",
            "Explanation",
            "Source",
            "Parent Terms",
            "Child Terms",
            "Equivalent Terms",
            "Equivalent Terms 2",
        ]
    )
    for row in rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {"A": 52, "B": 120, "C": 32, "D": 48, "E": 64, "F": 64, "G": 64}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge the AR6 working-group glossaries and AR7 SOD glossary."
    )
    parser.add_argument("--wgi", type=Path, default=DEFAULT_WGI_PATH)
    parser.add_argument("--wgii", type=Path, default=DEFAULT_WGII_PATH)
    parser.add_argument("--wgiii", type=Path, default=DEFAULT_WGIII_PATH)
    parser.add_argument("--ar7sod", type=Path, default=DEFAULT_AR7SOD_PATH)
    parser.add_argument("--sod-json", type=Path, default=DEFAULT_SOD_JSON_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_paths = {
        "AR6WGI": args.wgi,
        "AR6WGII": args.wgii,
        "AR6WGIII": args.wgiii,
        "AR7SOD": args.ar7sod,
        "SOD full text": args.sod_json,
    }
    missing_paths = [path for path in source_paths.values() if not path.is_file()]
    if missing_paths:
        raise FileNotFoundError(f"Glossary document not found: {missing_paths[0]}")

    entries_by_source = {
        source: parse_glossary(source_paths[source], source)
        for source in SOURCE_ORDER[:-1]
    }
    entries_by_source["AR7SOD"] = parse_glossary_workbook(
        source_paths["AR7SOD"], "AR7SOD"
    )
    glossary = merge_glossaries(entries_by_source)
    text_blocks = load_sod_text_blocks(args.sod_json)
    equivalent_terms_2 = scan_equivalent_terms_2(glossary, text_blocks)
    rows, unresolved = workbook_rows(glossary, equivalent_terms_2)
    if not rows:
        raise RuntimeError("No merged glossary entries were produced.")

    write_workbook(rows, args.output)
    source_counts = Counter(row[2] for row in rows)
    parsed_counts = ", ".join(
        f"{source}: {len(entries_by_source[source])}" for source in SOURCE_ORDER
    )
    merged_counts = ", ".join(
        f"{source}: {count}" for source, count in sorted(source_counts.items())
    )
    alternative_count = sum(len(alternatives) for alternatives in equivalent_terms_2.values())
    alternative_term_count = sum(bool(alternatives) for alternatives in equivalent_terms_2.values())
    print(
        f"Wrote {len(rows)} terms to {args.output} "
        f"(parsed {parsed_counts}; merged {merged_counts}; "
        f"unresolved parent links: {len(unresolved)}; "
        f"Equivalent Terms 2: {alternative_count} alternatives for "
        f"{alternative_term_count} terms)"
    )


if __name__ == "__main__":
    main()