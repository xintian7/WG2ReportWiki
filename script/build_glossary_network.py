#!/usr/bin/env python3
"""Build a Cypher glossary co-occurrence network for the SRCities reader."""

from __future__ import annotations

from collections import defaultdict
from itertools import combinations
import json
from pathlib import Path
import re

from openpyxl import load_workbook

from srcities_streamlit_app import (
    EXECUTIVE_SUMMARY_ITEM_RE,
    FIGURE_CAPTION_RE,
    FIRST_LEVEL_STATEMENT_RE,
    Glossary,
    STATEMENT_NUMBER_RE,
    glossary_usage_counts,
    markdown_to_plain_text,
    normalize_source_label,
    parse_markdown_sections,
)


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
REPORT_PATH = REPO_ROOT / "data" / "report" / "SRCities_FOD_SPM_Final.md"
GLOSSARY_PATHS = (
    REPO_ROOT / "data" / "Glossary" / "AR6FGD_Glossary.xlsx",
    REPO_ROOT / "data" / "Glossary" / "AR7SOD_Glossary.xlsx",
)
EXECUTIVE_SUMMARY_PATHS = {
    f"Chapter {chapter}": REPO_ROOT / "data" / "ES" / f"SRCities_SOD_Ch{chapter:02d}_Final_executive_summary.md"
    for chapter in range(1, 6)
}
OUTPUT_PATH = REPO_ROOT / "data" / "network" / "SRCities_glossary_network.cypher"

SENTENCE_BOUNDARY_RE = re.compile(r"(?<=[.!?])\s+")
ASSESSMENT_MARKUP_RE = re.compile(
    r"\((?:very |high |medium |low )?(?:confidence|agreement|evidence)(?:,\s*(?:very |high |medium |low )?(?:confidence|agreement|evidence))?\)|\{[^}]*\}",
    re.IGNORECASE,
)


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def cell_text(value: object) -> str:
    """Normalize a worksheet cell using the same empty-value behavior as the app."""
    return normalize_text(str(value)) if value is not None else ""


def load_glossary() -> Glossary:
    """Load and merge the glossary workbooks used by Glossary Overview."""
    glossary: Glossary = {}
    for path in GLOSSARY_PATHS:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        headers = [cell_text(cell.value).casefold() for cell in worksheet[1]]
        term_index = headers.index("term")
        explanation_index = headers.index("explanation")
        parent_index = headers.index("parent") if "parent" in headers else None
        source_index = headers.index("source")

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            term = cell_text(row[term_index] if term_index < len(row) else None)
            explanation = cell_text(row[explanation_index] if explanation_index < len(row) else None)
            parent = cell_text(row[parent_index]) if parent_index is not None and parent_index < len(row) else ""
            source = normalize_source_label(cell_text(row[source_index] if source_index < len(row) else None))
            if not term or not explanation or not source:
                continue

            definition = (term, explanation, parent, source)
            definitions = glossary.setdefault(term.casefold(), [])
            if definition not in definitions:
                definitions.append(definition)
        workbook.close()

    source_order = {"SRCities-SOD": 0, "AR6": 1}
    for definitions in glossary.values():
        definitions.sort(key=lambda definition: (source_order.get(definition[3], 2), definition[3]))
    return glossary


def load_executive_summaries() -> dict[str, str]:
    """Load the Executive Summary files used by Glossary Overview."""
    return {chapter: path.read_text(encoding="utf-8") for chapter, path in EXECUTIVE_SUMMARY_PATHS.items()}


def usage_sentence_records(
    report_text: str,
    executive_summaries: dict[str, str],
) -> list[tuple[str, str]]:
    """Return the same SPM and ES sentence units used for Glossary Overview counts."""
    _, sections = parse_markdown_sections(report_text)
    records: list[tuple[str, str]] = []

    for section_name, lines in sections.items():
        for line in lines:
            number_match = STATEMENT_NUMBER_RE.match(line)
            first_level_match = FIRST_LEVEL_STATEMENT_RE.match(line)
            figure_match = FIGURE_CAPTION_RE.match(line)
            if figure_match and section_name.startswith("Section "):
                statement_number = f"{section_name.removeprefix('Section ')}-Figure {figure_match.group(1)}"
                excerpts = [figure_match.group(2)]
            elif number_match:
                statement_number, statement_text = number_match.groups()
                excerpts = SENTENCE_BOUNDARY_RE.split(statement_text)
            elif first_level_match:
                statement_number, statement_text = first_level_match.groups()
                excerpts = SENTENCE_BOUNDARY_RE.split(statement_text)
            else:
                statement_number = section_name
                excerpts = SENTENCE_BOUNDARY_RE.split(line)

            records.extend((statement_number, sentence.strip()) for sentence in excerpts if sentence.strip())

    for summary_text in executive_summaries.values():
        for raw_line in summary_text.splitlines():
            item_match = EXECUTIVE_SUMMARY_ITEM_RE.match(raw_line.strip())
            if not item_match:
                continue
            statement_number, markdown_text = item_match.groups()
            records.extend(
                (statement_number, sentence.strip())
                for sentence in SENTENCE_BOUNDARY_RE.split(markdown_to_plain_text(markdown_text))
                if sentence.strip()
            )

    return records


def active_terms_in_text(
    text: str,
    active_terms_by_initial: dict[str, list[str]],
) -> set[str]:
    """Find active glossary terms in one sentence after removing assessment markup."""
    cleaned_text = ASSESSMENT_MARKUP_RE.sub("", text).casefold()
    matched_terms: set[str] = set()

    for start, character in enumerate(cleaned_text):
        if start and (cleaned_text[start - 1].isalnum() or cleaned_text[start - 1] == "_"):
            continue
        for term_key in active_terms_by_initial.get(character, []):
            end = start + len(term_key)
            if cleaned_text.startswith(term_key, start) and (
                end == len(cleaned_text) or not (cleaned_text[end].isalnum() or cleaned_text[end] == "_")
            ):
                matched_terms.add(term_key)
    return matched_terms


def cypher_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def build_network() -> tuple[list[dict[str, str | int]], list[dict[str, object]]]:
    """Build graph nodes from all active overview terms and sentence-level co-occurrences."""
    report_text = REPORT_PATH.read_text(encoding="utf-8")
    glossary = load_glossary()
    executive_summaries = load_executive_summaries()
    _, sections = parse_markdown_sections(report_text)
    usage_counts = glossary_usage_counts(glossary, sections, executive_summaries)
    active_terms = {term_key for term_key, count in usage_counts.items() if count > 0}
    active_terms_by_initial: dict[str, list[str]] = defaultdict(list)
    for term_key in active_terms:
        active_terms_by_initial[term_key[0]].append(term_key)
    edge_evidence: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)

    for statement_number, sentence in usage_sentence_records(report_text, executive_summaries):
        for source, target in combinations(sorted(active_terms_in_text(sentence, active_terms_by_initial)), 2):
            edge_evidence[(source, target)].append(
                {"statement": statement_number, "sentence": sentence}
            )

    nodes = [
        {"id": term_key, "label": glossary[term_key][0][0], "count": usage_counts[term_key]}
        for term_key in sorted(active_terms, key=lambda key: glossary[key][0][0].casefold())
    ]
    edges = [
        {
            "source": source,
            "target": target,
            "weight": len(evidence),
            "statements": sorted({item["statement"] for item in evidence}),
            "evidence": evidence,
        }
        for (source, target), evidence in sorted(edge_evidence.items())
    ]
    return nodes, edges


def write_cypher(nodes: list[dict[str, str | int]], edges: list[dict[str, object]]) -> None:
    """Write executable Cypher plus JSON metadata used by the Streamlit renderer."""
    lines = [
        "// SRCities glossary term co-occurrence network.",
        "// Nodes are all terms active in Glossary Overview across the SPM and Executive Summaries.",
        "// Relationships join terms appearing in the same sentence.",
        "",
    ]
    for node in nodes:
        lines.extend(
            [
                f"// NODE {json.dumps(node, ensure_ascii=False)}",
                f"MERGE (term:GlossaryTerm {{id: {cypher_string(str(node['id']))}}})",
                f"SET term.label = {cypher_string(str(node['label']))}, term.count = {node['count']};",
                "",
            ]
        )
    for edge in edges:
        lines.extend(
            [
                f"// EDGE {json.dumps(edge, ensure_ascii=False)}",
                f"MATCH (source:GlossaryTerm {{id: {cypher_string(str(edge['source']))}}}), "
                f"(target:GlossaryTerm {{id: {cypher_string(str(edge['target']))}}})",
                "MERGE (source)-[connection:CO_OCCURS]->(target)",
                f"SET connection.weight = {edge['weight']}, "
                f"connection.statements = {json.dumps(edge['statements'])}, "
                f"connection.evidence = {cypher_string(json.dumps(edge['evidence'], ensure_ascii=False))};",
                "",
            ]
        )
    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    nodes, edges = build_network()
    write_cypher(nodes, edges)
    print(f"Wrote {OUTPUT_PATH}: {len(nodes)} active terms, {len(edges)} co-occurrence links.")


if __name__ == "__main__":
    main()