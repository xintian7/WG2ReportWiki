#!/usr/bin/env python3
"""Organize the AR6 WGI FGD glossary into a four-column Excel workbook.

Direct cross-references of the form ``See Term`` are assigned the referenced term
as their parent. For ``See Term (under Parent)``, the explicitly named outer
parent is used. Incidental ``See also`` citations remain unassigned.

Example:
    /opt/anaconda3/envs/tsu/bin/python script/ar6_fgd_glossary_to_xlsx.py
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


SOURCE_NAME = "AR6"
DIRECT_PARENT_PREFIX_RE = re.compile(r"^See(?!\s+also\b)\s+.+?\s+\(under\s+", re.IGNORECASE)
DIRECT_SEE_RE = re.compile(r"^See(?!\s+also\b)\s+(?P<term>.+?)\.?$", re.IGNORECASE)


def normalize_text(value: object) -> str:
    """Convert worksheet values to normalized text."""
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def parent_from_explanation(explanation: str) -> str:
    """Extract a parent only from an unambiguous direct see-reference."""
    match = DIRECT_PARENT_PREFIX_RE.match(explanation)
    if match:
        parent_with_outer_closing = explanation[match.end() :].rstrip(".").rstrip()
        if parent_with_outer_closing.endswith(")"):
            return parent_with_outer_closing[:-1].rstrip()

    match = DIRECT_SEE_RE.fullmatch(explanation)
    if not match:
        return ""
    return match.group("term").rstrip(".").strip()


def read_entries(input_path: Path) -> list[tuple[str, str, str]]:
    """Read Term and Explanation rows from the source workbook."""
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [normalize_text(cell.value).casefold() for cell in worksheet[1]]
    try:
        term_index = headers.index("term")
        explanation_index = headers.index("explanation")
    except ValueError as error:
        raise ValueError("Source workbook must contain Term and Explanation columns.") from error

    entries: list[tuple[str, str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        term = normalize_text(row[term_index] if term_index < len(row) else "")
        explanation = normalize_text(row[explanation_index] if explanation_index < len(row) else "")
        if term:
            entries.append((term, explanation, parent_from_explanation(explanation)))

    return entries


def write_workbook(entries: list[tuple[str, str, str]], output_path: Path) -> None:
    """Write the organized glossary to a formatted workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Glossary"
    worksheet.append(["Term", "Explanation", "Parent", "Source"])
    for term, explanation, parent in entries:
        worksheet.append([term, explanation, parent, SOURCE_NAME])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for column in worksheet.iter_cols(min_row=2):
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.column_dimensions["A"].width = 52
    worksheet.column_dimensions["B"].width = 120
    worksheet.column_dimensions["C"].width = 52
    worksheet.column_dimensions["D"].width = 24
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Organize the AR6 WGI FGD glossary into Excel.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/Glossary/Orginal_IPCC_AR6_WGI_FGD_AnnexVII_Glossary.xlsx"),
        help="Path to the original AR6 WGI FGD glossary workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/Glossary/AR6FGD_Glossary.xlsx"),
        help="Path for the organized glossary workbook.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"Source workbook not found: {args.input}")

    entries = read_entries(args.input)
    if not entries:
        raise RuntimeError("No glossary entries were found in the source workbook.")

    write_workbook(entries, args.output)
    print(f"Wrote {len(entries)} glossary rows to {args.output}")


if __name__ == "__main__":
    main()